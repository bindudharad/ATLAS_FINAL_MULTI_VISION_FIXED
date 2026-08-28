"""Tests for the UIA-anchored data-entry flow.

Covers the new pieces without touching a real UI: the new state-machine
transitions, the source-pair OCR parser, the field-map JSON round-trip, the
upload-button picker, and the end-to-end AgentLoop run against a synthetic
``UiaFieldMap`` (exact form geometry injected into an otherwise empty scene).
"""

from __future__ import annotations

from pathlib import Path

import numpy as np

from atlas.act.controls import ControlInterface, ControlOutcome
from atlas.act.executor import ActionExecutor
from atlas.core.states import AgentState, StateMachine
from atlas.core.events import EventType, get_event_bus
from atlas.mapping.mapper import SemanticMapper
from atlas.mapping.uia_map import UiaFieldMap, UiaFieldMapBuilder, build_hybrid_mappings, pair_source_pairs
from atlas.observe.uia import UiaNode
from atlas.reason.planner import ActionPlanner
from atlas.reason.recovery import RecoveryPlanner
from atlas.target.base import TargetAdapter, TargetInfo
from atlas.understanding.source import SourceReader, SourceRecord
from atlas.vision.models import BBox, ElementType, OcrText, SceneDescription
from atlas.vision.scene import SceneAnalysis
from atlas.workflow.loop import AgentLoop

from tests.test_mpf_integration import PassVerifier, RecordingControls, StubKeyboard, StubMouse

from plugins.mpf.plugin import MpfPlugin


# ---------------------------------------------------------------------------
# New state machine
# ---------------------------------------------------------------------------


def test_state_machine_new_states() -> None:
    sm = StateMachine()
    sm.force(AgentState.ATTACHING)
    sm.transition(AgentState.WAITING_FOR_START_FIELD)
    assert sm.state == AgentState.WAITING_FOR_START_FIELD
    sm.transition(AgentState.FIELD_MAPPING)
    assert sm.state == AgentState.FIELD_MAPPING
    sm.transition(AgentState.WATCHING)
    assert sm.state == AgentState.WATCHING


def test_state_machine_new_states_active() -> None:
    sm = StateMachine()
    for state in (AgentState.WAITING_FOR_START_FIELD, AgentState.FIELD_MAPPING):
        sm.force(state)
        assert sm.is_active()


# ---------------------------------------------------------------------------
# Source pair parsing
# ---------------------------------------------------------------------------


def _ocr(texts: list[str]) -> list[OcrText]:
    return [OcrText(text=t, bbox=BBox(0, i * 20, 300, 16)) for i, t in enumerate(texts)]


def test_pair_source_pairs_colon_lines() -> None:
    lines = _ocr([
        "Application Number : MPF-100",
        "Full Name : KRISHNA",
        "DOB : 21 March 1996",
    ])
    pairs = pair_source_pairs(lines, [])
    assert dict(pairs) == {
        "Application Number": "MPF-100",
        "Full Name": "KRISHNA",
        "DOB": "21 March 1996",
    }


def test_pair_source_pairs_uia_labels_fill_in() -> None:
    lines = _ocr(["Application Number : MPF-100"])
    labels = [UiaNode(name="Full Name", control_type="Text"), UiaNode(name="Application Number", control_type="Text")]
    pairs = dict(pair_source_pairs(lines, labels))
    assert pairs["Application Number"] == "MPF-100"
    assert "Full Name" in pairs  # known label from UIA, no OCR value


def test_pair_source_pairs_geometric_two_column() -> None:
    """A row of two sibling UIA text nodes pairs label -> value.

    Mirrors the real MPF left panel where each field's label and value are
    separate text controls on the same row (no OCR colon line).
    """
    labels = [
        UiaNode(name="App No", control_type="Text", rect=BBox(539, 334, 41, 14)),
        UiaNode(name="55096808", control_type="Text", rect=BBox(637, 334, 51, 14)),
        UiaNode(name="Full Name", control_type="Text", rect=BBox(539, 366, 54, 15)),
        UiaNode(name="RAJESH KUMAR", control_type="Text", rect=BBox(637, 366, 97, 15)),
    ]
    pairs = dict(pair_source_pairs([], labels))
    assert pairs["App No"] == "55096808"
    assert pairs["Full Name"] == "RAJESH KUMAR"


def test_pair_source_pairs_geometric_skips_wide_header() -> None:
    """Section headers and cross-column duplicate labels must not pair.

    Note: section headers are also excluded from the pairs dict entirely
    (not kept with an empty value) - see ``is_noise_label`` in
    ``atlas/mapping/uia_map.py``. A header was never a real source field and
    keeping it (even as "") inflated the unmapped-label count and dragged
    down reported source coverage in real runs.
    """
    labels = [
        UiaNode(name="Member Basic Information", control_type="Text", rect=BBox(539, 315, 253, 16)),
        UiaNode(name="App No", control_type="Text", rect=BBox(539, 334, 41, 14)),
        UiaNode(name="55096808", control_type="Text", rect=BBox(637, 334, 51, 14)),
        # duplicate label from the second column - different row, no value
        UiaNode(name="App No", control_type="Text", rect=BBox(841, 344, 45, 16)),
    ]
    pairs = dict(pair_source_pairs([], labels))
    assert pairs["App No"] == "55096808"
    assert "Member Basic Information" not in pairs


def test_pair_source_pairs_geometric_wide_header_not_paired_as_label() -> None:
    """A wide header next to a duplicate label on the same row must not pair
    the header with the duplicate as if it were its value, and neither must
    ever surface as a source field (see is_noise_label / NOISE_SECTION_HEADERS)."""
    labels = [
        UiaNode(name="Religious and Astro Information", control_type="Text", rect=BBox(539, 524, 253, 16)),
        UiaNode(name="State", control_type="Text", rect=BBox(841, 525, 41, 15)),
    ]
    pairs = dict(pair_source_pairs([], labels))
    assert "Religious and Astro Information" not in pairs
    assert pairs["State"] == ""


# ---------------------------------------------------------------------------
# Field-first mapping regressions (source panel label/value siblings)
# ---------------------------------------------------------------------------


def _parent_group(name: str) -> dict[str, str]:
    return {"name": name, "control_type": "Group", "automation_id": ""}


def test_pair_source_pairs_empty_colon_line_does_not_block_geometric_pair() -> None:
    """An OCR 'Mother Status:' line with an empty value must not stop the
    geometric pairing that recovers the real value ('Alive')."""
    lines = _ocr(["Mother Status:"])
    labels = [
        UiaNode(name="Mother Status", control_type="Text", rect=BBox(539, 676, 71, 14), parent=_parent_group("Record summary")),
        UiaNode(name="Alive", control_type="Text", rect=BBox(637, 676, 29, 14), parent=_parent_group("Record summary")),
    ]
    pairs = dict(pair_source_pairs(lines, labels))
    assert pairs["Mother Status"] == "Alive"


def test_pair_source_pairs_dominant_parent_row_filter() -> None:
    """A stray menu fragment on the State row must not steal the value; the
    label/value siblings share one parent group."""
    labels = [
        UiaNode(name="State", control_type="Text", rect=BBox(539, 228, 26, 15), parent=_parent_group("Record summary")),
        UiaNode(name="ecord", control_type="Text", rect=BBox(585, 226, 34, 18), parent=_parent_group("Application menu")),
        UiaNode(name="Tamil Nadu", control_type="Text", rect=BBox(637, 228, 61, 15), parent=_parent_group("Record summary")),
    ]
    pairs = dict(pair_source_pairs([], labels))
    assert pairs["State"] == "Tamil Nadu"


def test_build_hybrid_mappings_field_first_no_value_theft() -> None:
    """Right fields claim their own labels; source VALUES (numeric IDs, names,
    dates) and wide section headers must never become mapping sources."""
    labels = [
        UiaNode(name="Member Basic Information", control_type="Text", rect=BBox(539, 112, 253, 16), parent=_parent_group("Record summary")),
        UiaNode(name="App No", control_type="Text", rect=BBox(539, 131, 41, 14), parent=_parent_group("Record summary")),
        UiaNode(name="88616739", control_type="Text", rect=BBox(637, 131, 51, 14), parent=_parent_group("Record summary")),
        UiaNode(name="MBI Code", control_type="Text", rect=BBox(539, 147, 54, 15), parent=_parent_group("Record summary")),
        UiaNode(name="MBI2102426874", control_type="Text", rect=BBox(637, 147, 88, 15), parent=_parent_group("Record summary")),
        UiaNode(name="Full Name", control_type="Text", rect=BBox(539, 163, 54, 15), parent=_parent_group("Record summary")),
        UiaNode(name="ANITA SHARMA", control_type="Text", rect=BBox(637, 163, 95, 15), parent=_parent_group("Record summary")),
        UiaNode(name="Date Of Birth", control_type="Text", rect=BBox(539, 196, 70, 14), parent=_parent_group("Record summary")),
        UiaNode(name="02 February 1996", control_type="Text", rect=BBox(637, 196, 90, 14), parent=_parent_group("Record summary")),
        UiaNode(name="District", control_type="Text", rect=BBox(539, 244, 38, 15), parent=_parent_group("Record summary")),
        UiaNode(name="Vijayawada", control_type="Text", rect=BBox(637, 244, 60, 15), parent=_parent_group("Record summary")),
    ]
    fields = [
        UiaNode(name="App No", control_type="Edit", automation_id="applicationNumber", rect=BBox(972, 325, 257, 26)),
        UiaNode(name="MBI Code", control_type="Edit", automation_id="mbiCode", rect=BBox(972, 354, 257, 25)),
        UiaNode(name="Full Name", control_type="Edit", automation_id="fullName", rect=BBox(972, 382, 257, 26)),
        UiaNode(name="District", control_type="ComboBox", automation_id="district", rect=BBox(972, 526, 257, 26)),
    ]
    mappings = build_hybrid_mappings(labels, fields)
    by_target = {m["target"]: m for m in mappings}
    assert by_target["App No"]["source"] == "App No"
    assert by_target["MBI Code"]["source"] == "MBI Code"
    assert by_target["Full Name"]["source"] == "Full Name"
    assert by_target["District"]["source"] == "District"
    sources = {m["source"] for m in mappings}
    assert "88616739" not in sources
    assert "MBI2102426874" not in sources
    assert "ANITA SHARMA" not in sources
    assert "02 February 1996" not in sources
    assert "Date Of Birth" not in sources  # no right field named Date Of Birth
    assert "Member Basic Information" not in sources  # section header not mapped


def test_build_hybrid_mappings_prefers_exact_name_over_alias() -> None:
    """Field 'App No' must pick the exact label 'App No', not the alias
    'Application No.'."""
    labels = [
        UiaNode(name="Application No.", control_type="Text", rect=BBox(529, 260, 92, 18), parent=_parent_group("Record commands")),
        UiaNode(name="App No", control_type="Text", rect=BBox(539, 131, 41, 14), parent=_parent_group("Record summary")),
    ]
    fields = [UiaNode(name="App No", control_type="Edit", automation_id="applicationNumber", rect=BBox(972, 325, 257, 26))]
    mappings = build_hybrid_mappings(labels, fields)
    assert len(mappings) == 1
    assert mappings[0]["source"] == "App No"
    assert mappings[0]["method"] == "semantic"


def test_build_hybrid_mappings_does_not_map_do_b_to_wrong_field() -> None:
    """The DOB source label must not cascade onto a non-date field like
    District."""
    labels = [
        UiaNode(name="Date Of Birth", control_type="Text", rect=BBox(539, 196, 70, 14), parent=_parent_group("Record summary")),
        UiaNode(name="02 February 1996", control_type="Text", rect=BBox(637, 196, 90, 14), parent=_parent_group("Record summary")),
        UiaNode(name="District", control_type="Text", rect=BBox(539, 244, 38, 15), parent=_parent_group("Record summary")),
        UiaNode(name="Vijayawada", control_type="Text", rect=BBox(637, 244, 60, 15), parent=_parent_group("Record summary")),
    ]
    fields = [UiaNode(name="District", control_type="ComboBox", automation_id="district", rect=BBox(972, 526, 257, 26))]
    mappings = build_hybrid_mappings(labels, fields)
    assert [m["source"] for m in mappings] == ["District"]


def test_field_map_load_preserves_parent_groups() -> None:
    """Serialising/loading a field map must keep each label's parent group so
    the same-parent source-value exclusion still works offline."""
    node = UiaNode(name="State", control_type="Text", rect=BBox(539, 228, 26, 15), parent=_parent_group("Record summary"))
    m = UiaFieldMap(left_labels=[node], right_fields=[], mappings=[], client_origin=(0, 29))
    loaded = UiaFieldMap.from_dict(m.to_dict())
    assert loaded.left_labels[0].parent == _parent_group("Record summary")


# ---------------------------------------------------------------------------
# Field map serialization
# ---------------------------------------------------------------------------


def _sample_field_map() -> UiaFieldMap:
    start = UiaNode(name="Full Name", control_type="Edit", handle=1001, rect=BBox(500, 40, 200, 24))
    left = [
        UiaNode(name="Application Number", control_type="Text", rect=BBox(20, 20, 150, 18)),
        UiaNode(name="Full Name", control_type="Text", rect=BBox(20, 50, 150, 18)),
    ]
    right = [
        UiaNode(name="Full Name", control_type="Edit", handle=2001, rect=BBox(500, 40, 200, 24)),
        UiaNode(name="Gender", control_type="ComboBox", handle=2002, rect=BBox(500, 80, 200, 24), options=["Male", "Female"]),
        UiaNode(name="Date Of Birth", control_type="Edit", handle=2003, rect=BBox(500, 120, 200, 24),
                type_override=ElementType.DATE_PICKER),
    ]
    upload = UiaNode(name="Upload Details", control_type="Button", handle=3001, rect=BBox(500, 300, 140, 32))
    return UiaFieldMap(
        start_control=start,
        left_labels=left,
        right_fields=right,
        upload_button=upload,
        left_rect=BBox(10, 10, 300, 200),
        right_rect=BBox(480, 30, 300, 320),
        mappings=[{"source": "Full Name", "target": "Full Name", "confidence": 0.98}],
        client_origin=(0, 0),
        client_size=(1024, 768),
    )


def test_field_map_json_round_trip(tmp_path: Path) -> None:
    field_map = _sample_field_map()
    path = tmp_path / "field_map.json"
    field_map.save(path)
    loaded = UiaFieldMap.load(path)
    assert loaded is not None
    assert loaded.start_control is not None and loaded.start_control.name == "Full Name"
    assert len(loaded.left_labels) == 2
    assert len(loaded.right_fields) == 3
    assert loaded.right_fields[1].options == ["Male", "Female"]
    assert loaded.upload_button is not None and loaded.upload_button.name == "Upload Details"
    assert loaded.left_rect == BBox(10, 10, 300, 200)


def test_builder_attaches_declared_type_and_options() -> None:
    builder = UiaFieldMapBuilder(declared_fields={
        "gender": {"type": "combobox", "options": ["Male", "Female"]},
        "date of birth": {"type": "date_picker"},
    })
    node = UiaNode(name="Gender", control_type="Edit", handle=5)
    node = builder._attach_declared(node)
    assert node.type_override == ElementType.COMBOBOX
    assert node.options == ["Male", "Female"]

    dob = builder._attach_declared(UiaNode(name="Date Of Birth", control_type="Edit"))
    assert dob.type_override == ElementType.DATE_PICKER


def test_upload_button_picker_word_boundary() -> None:
    buttons = [
        UiaNode(name="Blue Book DSA", control_type="Button", rect=BBox(180, 700, 100, 24)),
        UiaNode(name="Upload Details", control_type="Button", rect=BBox(500, 300, 140, 32)),
        UiaNode(name="OK", control_type="Button", rect=BBox(10, 10, 40, 20)),
    ]
    picked = UiaFieldMapBuilder._pick_upload_button(buttons)
    assert picked is not None and picked.name == "Upload Details"


class _NoScrollBackend:
    """Fake UIA backend that records whether scroll_into_view was requested."""

    def __init__(self, fields: list[UiaNode], client_size: tuple[int, int] = (1024, 768)) -> None:
        self._fields = fields
        self._size = client_size
        self.scroll_calls = 0

    def client_origin(self, hwnd: int) -> tuple[int, int]:
        return (0, 0)

    def client_size(self, hwnd: int) -> tuple[int, int]:
        return self._size

    def editable_fields(self, hwnd: int) -> list[UiaNode]:
        return self._fields

    def text_nodes(self, hwnd: int) -> list[UiaNode]:
        return []

    def buttons(self, hwnd: int) -> list[UiaNode]:
        return []

    def scroll_into_view(self, node: UiaNode) -> UiaNode:
        self.scroll_calls += 1
        return node


def test_field_map_build_never_scrolls() -> None:
    """Building the UIA field map must never scroll, even for below-fold fields.

    This was the premature-scroll root cause: right after attach the builder
    scrolled every field below the fold. The map is built from the tree as-is;
    the workflow's reveal pass scrolls per-viewport, gated by can_scroll()."""
    fields = [
        UiaNode(name="App No", control_type="Edit", rect=BBox(500, 300, 200, 24), enabled=True),
        UiaNode(name="Full Name", control_type="Edit", rect=BBox(500, 700, 200, 24), enabled=True),
        UiaNode(name="Gender", control_type="ComboBox", rect=BBox(500, 1400, 200, 24), enabled=True),
    ]
    backend = _NoScrollBackend(fields, client_size=(1024, 768))
    builder = UiaFieldMapBuilder(backend=backend)
    field_map = builder.build(2000)
    assert field_map.has_form
    assert len(field_map.right_fields) == 3
    assert backend.scroll_calls == 0


class _ParentGroupBackend(_NoScrollBackend):
    """Fake backend returning both editables and source/form label text nodes."""

    def __init__(self, fields: list[UiaNode], text: list[UiaNode], client_size=(1024, 768)) -> None:
        super().__init__(fields, client_size=client_size)
        self._text = text

    def text_nodes(self, hwnd: int) -> list[UiaNode]:
        return self._text


def test_field_map_build_excludes_form_labels_from_left_labels() -> None:
    """Labels inside a right-form parent group must never join ``left_labels``.

    The MPF right column starts around x=840 - left of the window mid-line
    (960) - so its form labels ("Gender Marital Status", "Religious and Astro
    Information") were leaking into the source label pool and inflating
    ``left_rect`` across the divider. They share their parent Group with the
    editable right fields; source-panel labels live under a pure-text group."""
    fields = [
        UiaNode(name="gender", control_type="ComboBox", rect=BBox(961, 382, 54, 26),
                parent={"name": "Member Basic Information", "control_type": "Group"}),
        UiaNode(name="nakshatra", control_type="ComboBox", rect=BBox(961, 700, 132, 26),
                parent={"name": "Religious and Astro Information", "control_type": "Group"}),
    ]
    text = [
        UiaNode(name="Member Name", control_type="Text", rect=BBox(585, 186, 80, 22),
                parent={"name": "Record summary", "control_type": "Group"}),
        UiaNode(name="Application No", control_type="Text", rect=BBox(585, 230, 140, 22),
                parent={"name": "Record summary", "control_type": "Group"}),
        UiaNode(name="Gender Marital Status", control_type="Text", rect=BBox(961, 382, 160, 22),
                parent={"name": "Member Basic Information", "control_type": "Group"}),
        UiaNode(name="Religious and Astro Information", control_type="Text", rect=BBox(961, 666, 200, 22),
                parent={"name": "Religious and Astro Information", "control_type": "Group"}),
    ]
    backend = _ParentGroupBackend(fields, text, client_size=(1920, 991))
    field_map = UiaFieldMapBuilder(backend=backend).build(2000)

    label_names = {n.name for n in field_map.left_labels}
    assert "Member Name" in label_names
    assert "Application No" in label_names
    assert "Gender Marital Status" not in label_names
    assert "Religious and Astro Information" not in label_names

    left_rect = field_map.left_rect
    assert left_rect is not None
    assert left_rect.right <= 800, f"left_rect leaks into the right form: {left_rect}"


# ---------------------------------------------------------------------------
# End-to-end loop against a synthetic field map
# ---------------------------------------------------------------------------


class FieldMapFakeTarget(TargetAdapter):
    name = "fake-fieldmap"

    def __init__(self, records: list[dict]) -> None:
        self._records = records
        self._idx = 0
        self._info = TargetInfo(name="fake-fieldmap", title="MPF (Download and Upload Form)")

    def attach(self, hint: str | None = None) -> TargetInfo:
        return self._info

    def detach(self) -> None:
        pass

    def observe(self) -> SceneAnalysis | None:
        if self._idx < len(self._records):
            self._idx += 1
            return SceneAnalysis(scene=SceneDescription(
                window_title="MPF (Download and Upload Form)",
                layout_summary="empty scene",
                screen_offset=(0, 0),
            ))
        return None

    def is_alive(self) -> bool:
        return True

    def read_field_value(self, field_id: str) -> str | None:
        return None

    def current(self) -> dict:
        return self._records[min(max(self._idx - 1, 0), len(self._records) - 1)]


def _run_anchored_records(tmp_path: Path) -> tuple[AgentLoop, RecordingControls, dict]:
    get_event_bus().clear()  # event bus is a process singleton; drop prior history
    records = [
        {"key": "MPF-100", "name": "KRISHNA", "gender": "Male", "dob": "21 March 1996"},
        {"key": "MPF-200", "name": "RAVI KUMAR", "gender": "Female", "dob": "05 August 1990"},
        {"key": "MPF-300", "name": "SITA DEVI", "gender": "Male", "dob": "14 December 1988"},
    ]
    target = FieldMapFakeTarget(records)
    controls = RecordingControls()

    def ocr_callback(bbox: BBox) -> list[OcrText]:
        rec = target.current()
        return _ocr([
            f"Application Number : {rec['key']}",
            f"Full Name : {rec['name']}",
            f"Gender : {rec['gender']}",
            f"DOB : {rec['dob']}",
        ])

    plugin = MpfPlugin()
    mapper = SemanticMapper()
    for variant, canonical in plugin._config.get("aliases", {}).items():
        mapper.aliases.learn(variant, canonical)

    executor = ActionExecutor(
        mouse=StubMouse(), keyboard=StubKeyboard(), controls=controls,
        verifier=PassVerifier(), recovery=RecoveryPlanner(),
        verify_after_action=True, max_retries=3, retry_delay=0.0,
    )

    field_map = _sample_field_map()
    loop = AgentLoop(
        target=target, source_reader=SourceReader(), mapper=mapper,
        planner=ActionPlanner(verify_after_action=True), executor=executor,
        max_records=3, next_record_timeout=2.0, next_record_poll=0.05,
        scene_hook=plugin.refine_scene,
        field_map=field_map,
        ocr_callback=ocr_callback,
        debug_dir=tmp_path,
    )
    summary = loop.run()
    return loop, controls, summary.to_dict()


def test_anchored_loop_three_records(tmp_path: Path) -> None:
    loop, controls, summary = _run_anchored_records(tmp_path)

    assert summary["completed"] == 3, summary
    assert summary["failed"] == 0
    assert controls.typed == ["KRISHNA", "RAVI KUMAR", "SITA DEVI"]
    assert controls.selected == ["Male", "Female", "Male"]
    assert controls.dates == ["21 March 1996", "05 August 1990", "14 December 1988"]
    assert loop.state.value == "stopped"

    uploads = get_event_bus().history(EventType.UPLOAD_COMPLETED)
    assert len(uploads) == 3


def _field_map_for(rec: dict) -> UiaFieldMap:
    """A synthetic MPF field map whose LEFT source panel mirrors ``rec``.

    Each source field is a label text node with its value as a sibling text
    node on the same row, exactly like the real MPF left panel, so
    ``pair_source_pairs`` recovers label -> value from UIA alone (no OCR).
    Labels stay narrow (<= 120px) so they are not mistaken for section headers.
    """
    left = [
        UiaNode(name="Application Number", control_type="Text", rect=BBox(20, 20, 80, 18), handle=1101),
        UiaNode(name=rec["key"], control_type="Text", rect=BBox(120, 20, 130, 18), handle=1102),
        UiaNode(name="Full Name", control_type="Text", rect=BBox(20, 50, 80, 18), handle=1103),
        UiaNode(name=rec["name"], control_type="Text", rect=BBox(120, 50, 130, 18), handle=1104),
        UiaNode(name="Gender", control_type="Text", rect=BBox(20, 80, 80, 18), handle=1105),
        UiaNode(name=rec["gender"], control_type="Text", rect=BBox(120, 80, 130, 18), handle=1106),
        UiaNode(name="Date Of Birth", control_type="Text", rect=BBox(20, 110, 80, 18), handle=1107),
        UiaNode(name=rec["dob"], control_type="Text", rect=BBox(120, 110, 130, 18), handle=1108),
    ]
    right = [
        UiaNode(name="Application Number", control_type="Edit", handle=2000, rect=BBox(500, 10, 200, 24)),
        UiaNode(name="Full Name", control_type="Edit", handle=2001, rect=BBox(500, 40, 200, 24)),
        UiaNode(name="Gender", control_type="ComboBox", handle=2002, rect=BBox(500, 80, 200, 24),
                options=["Male", "Female"]),
        UiaNode(name="Date Of Birth", control_type="Edit", handle=2003, rect=BBox(500, 120, 200, 24),
                type_override=ElementType.DATE_PICKER),
    ]
    upload = UiaNode(name="Upload Details", control_type="Button", handle=3001, rect=BBox(500, 300, 140, 32))
    return UiaFieldMap(
        start_control=right[0],
        left_labels=left,
        right_fields=right,
        upload_button=upload,
        left_rect=BBox(10, 10, 300, 200),
        right_rect=BBox(480, 30, 300, 320),
        mappings=[
            {"source": "Full Name", "target": "Full Name", "confidence": 0.98},
            {"source": "Gender", "target": "Gender", "confidence": 0.98},
            {"source": "Date Of Birth", "target": "Date Of Birth", "confidence": 0.98},
        ],
        client_origin=(0, 0),
        client_size=(1024, 768),
    )


def test_field_driven_loop_multi_record_reset_detection(tmp_path: Path) -> None:
    """The field-driven loop processes 2 records back-to-back on MPF.

    After the upload click the simulated app advances its LEFT source panel to
    the next record. The UIA-first submit verification and the next-record
    await must detect that reset WITHOUT a VLM observe, and the second record
    must start clean (no stale field map / queue / data leaking in).
    """
    get_event_bus().clear()
    records = [
        {"key": "MPF-100", "name": "KRISHNA", "gender": "Male", "dob": "21 March 1996"},
        {"key": "MPF-200", "name": "RAVI KUMAR", "gender": "Female", "dob": "05 August 1990"},
        # The form advances to a third record after the final upload; the loop
        # must stop at max_records=2 without ever touching it.
        {"key": "MPF-300", "name": "SITA DEVI", "gender": "Male", "dob": "14 December 1988"},
    ]
    target = FieldMapFakeTarget(records)
    state = {"idx": 0}

    def refresh() -> UiaFieldMap:
        return _field_map_for(records[min(state["idx"], len(records) - 1)])

    class SubmitAwareControls(RecordingControls):
        def __init__(self) -> None:
            super().__init__()
            self.uploads = 0

        def click_field(self, bbox, field_id=None):
            if field_id and "uia-btn-3001" in str(field_id):
                state["idx"] = min(state["idx"] + 1, len(records) - 1)
                self.uploads += 1
            return super().click_field(bbox, field_id=field_id)

    controls = SubmitAwareControls()
    plugin = MpfPlugin()
    mapper = SemanticMapper()
    for variant, canonical in plugin._config.get("aliases", {}).items():
        mapper.aliases.learn(variant, canonical)

    executor = ActionExecutor(
        mouse=StubMouse(), keyboard=StubKeyboard(), controls=controls,
        verifier=PassVerifier(), recovery=RecoveryPlanner(),
        verify_after_action=True, max_retries=3, retry_delay=0.0,
    )
    loop = AgentLoop(
        target=target, source_reader=SourceReader(), mapper=mapper,
        planner=ActionPlanner(verify_after_action=True), executor=executor,
        max_records=2, next_record_timeout=2.0, next_record_poll=0.05,
        field_map=refresh(),
        field_map_refresh=refresh,
        field_driven=True,
        debug_dir=tmp_path,
    )
    summary = loop.run()

    assert summary.completed == 2, summary.to_dict()
    assert summary.failed == 0
    assert len(summary.records) == 2
    assert [r.record.record_key for r in summary.records] == ["MPF-100", "MPF-200"]
    # Each record was filled exactly once with ITS OWN data (no stale leak).
    # In the field-driven path the date picker is written via a SELECT action.
    # Application Number is now a real right-form field (see _field_map_for),
    # so it is typed too - each record types its own App No + Full Name only.
    assert controls.typed == ["MPF-100", "KRISHNA", "MPF-200", "RAVI KUMAR"]
    assert controls.selected == ["Male", "21 March 1996", "Female", "05 August 1990"]
    assert controls.uploads == 2
    assert loop.state.value == "stopped"


def test_anchored_loop_writes_debug_artifacts(tmp_path: Path) -> None:
    _run_anchored_records(tmp_path)
    for name in ("planner.json", "execution_plan.json", "execution.json", "verification.json"):
        assert (tmp_path / name).exists(), f"missing {name}"
    # Per-record session artifacts, including the extracted record.
    for name in ("record.json", "timeline.json"):
        assert (tmp_path / "session" / name).exists(), f"missing session/{name}"
    import json

    session = json.loads((tmp_path / "session" / "record.json").read_text(encoding="utf-8"))
    assert session["key"] == "MPF-300"
    # No failures -> no failure.json for a clean run.
    assert not (tmp_path / "failure.json").exists()


def test_anchored_loop_writes_failure_json_on_empty_source(tmp_path: Path) -> None:
    """With OCR finding nothing, the loop never self-terminates; it reports a
    no-record condition, keeps waiting, and only a user stop ends the run and
    writes failure.json."""
    import threading
    import time

    target = FieldMapFakeTarget([{"key": "MPF-100", "name": "KRISHNA", "gender": "Male", "dob": "21 March 1996"}])
    controls = RecordingControls()

    def ocr_callback(bbox: BBox) -> list[OcrText]:
        return []  # OCR finds nothing -> no source record

    plugin = MpfPlugin()
    mapper = SemanticMapper()
    for variant, canonical in plugin._config.get("aliases", {}).items():
        mapper.aliases.learn(variant, canonical)

    executor = ActionExecutor(
        mouse=StubMouse(), keyboard=StubKeyboard(), controls=controls,
        verifier=PassVerifier(), recovery=RecoveryPlanner(),
        verify_after_action=True, max_retries=3, retry_delay=0.0,
    )
    loop = AgentLoop(
        target=target, source_reader=SourceReader(), mapper=mapper,
        planner=ActionPlanner(verify_after_action=True), executor=executor,
        max_records=0, next_record_timeout=0.3, next_record_poll=0.02,
        scene_hook=plugin.refine_scene,
        field_map=_sample_field_map(),
        ocr_callback=ocr_callback,
        debug_dir=tmp_path,
    )

    def _stop() -> None:
        time.sleep(0.4)
        loop.stop()

    stopper = threading.Thread(target=_stop, daemon=True)
    stopper.start()
    summary = loop.run()
    stopper.join()

    assert summary.to_dict()["records"] == []
    no_record = tmp_path / "no_record.json"
    assert no_record.exists()
    failure = tmp_path / "failure.json"
    assert failure.exists()
    import json

    payload = json.loads(failure.read_text(encoding="utf-8"))
    assert payload["stopped_reason"] == "stopped by user"
    assert payload["no_record_reason"]
    # The NO_RECORD event surfaced the condition (no silent records=0).
    assert len(get_event_bus().history(EventType.NO_RECORD)) > 0


def test_auto_build_records_from_ocr_without_vlm(tmp_path: Path) -> None:
    """The auto (non-anchored) path: a start_control-free field map + OCR
    source pairs yield a real Record even though the VLM scene is empty."""
    from atlas.core.record_builder import RecordBuilder

    target = FieldMapFakeTarget([
        {"key": "MPF-900", "name": "ANIL", "gender": "Female", "dob": "01 January 1999"},
    ])
    controls = RecordingControls()
    get_event_bus().clear()

    def ocr_callback(bbox: BBox) -> list[OcrText]:
        rec = target.current()
        return _ocr([
            f"Application Number : {rec['key']}",
            f"Full Name : {rec['name']}",
            f"Gender : {rec['gender']}",
            f"DOB : {rec['dob']}",
        ])

    plugin = MpfPlugin()
    mapper = SemanticMapper()
    for variant, canonical in plugin._config.get("aliases", {}).items():
        mapper.aliases.learn(variant, canonical)

    declared = plugin._config.get("fields", {})
    record_builder = RecordBuilder(declared_fields=declared, aliases=plugin._config.get("aliases", {}))

    executor = ActionExecutor(
        mouse=StubMouse(), keyboard=StubKeyboard(), controls=controls,
        verifier=PassVerifier(), recovery=RecoveryPlanner(),
        verify_after_action=True, max_retries=3, retry_delay=0.0,
    )
    loop = AgentLoop(
        target=target, source_reader=SourceReader(), mapper=mapper,
        planner=ActionPlanner(verify_after_action=True), executor=executor,
        max_records=1, next_record_timeout=2.0, next_record_poll=0.05,
        scene_hook=plugin.refine_scene,
        field_map=_sample_field_map(),
        ocr_callback=ocr_callback,
        debug_dir=tmp_path,
        session_dir=tmp_path / "session",
        record_builder=record_builder,
    )
    summary = loop.run()

    assert summary.completed == 1
    record = summary.records[0].record
    assert record.pairs["Application Number"] == "MPF-900"
    assert record.pairs["Full Name"] == "ANIL"
    assert record.pairs["Gender"] == "Female"
    session = tmp_path / "session" / "record.json"
    assert session.exists()
    import json

    payload = json.loads(session.read_text(encoding="utf-8"))
    assert payload["key"] == "MPF-900"


# ---------------------------------------------------------------------------
# Assistant orchestration (fakes only - no live window)
# ---------------------------------------------------------------------------


class _FakeListener:
    """Stand-in for MouseClickListener: returns a fixed click point."""

    def __init__(self, click: tuple[int, int]) -> None:
        self._click = click

    def wait_for_click(self, timeout: float = 0.0) -> tuple[int, int] | None:
        return self._click

    def stop(self) -> None:
        pass


class _FakeBackend:
    """Stand-in for UiaBackend: resolves the click to an editable node."""

    def __init__(self, node: UiaNode) -> None:
        self._node = node

    def client_origin(self, hwnd: int) -> tuple[int, int]:
        return (0, 29)

    def client_size(self, hwnd: int) -> tuple[int, int]:
        return (800, 600)

    def element_at(self, x: int, y: int) -> UiaNode:
        return self._node

    def focused(self) -> UiaNode:
        return self._node

    def editable_fields(self, hwnd: int) -> list[UiaNode]:
        return [self._node]

    def text_nodes(self, hwnd: int) -> list[UiaNode]:
        return []

    def buttons(self, hwnd: int) -> list[UiaNode]:
        return []

    def scroll_into_view(self, node: UiaNode) -> UiaNode:
        return node

    def dump_tree(self, hwnd: int) -> list[dict]:
        return []


def test_assistant_captures_start_control_and_builds_map(tmp_path: Path, monkeypatch) -> None:
    from atlas.assistant.assistant import Assistant
    from atlas.observe.uia import UiaBackend
    from atlas.target.desktop import DesktopTarget
    from atlas.vision.capture import ScreenGrabber

    start = UiaNode(
        name="Full Name", control_type="Edit", handle=2001,
        rect=BBox(500, 40, 200, 24), enabled=True,
    )
    backend = _FakeBackend(start)
    monkeypatch.setattr(UiaBackend, "_instance", backend)
    monkeypatch.setattr(
        UiaBackend,
        "instance",
        classmethod(lambda cls: backend),
    )

    class _FakeTarget(DesktopTarget):
        class _Info:
            handle = 2000
            title = "MPF (Download and Upload Form)"

        def __init__(self) -> None:
            self.__dict__["_info"] = self._Info()

    class _FakeGrabber(ScreenGrabber):
        def grab_rect(self, x, y, width, height):  # noqa: ARG002
            from PIL import Image
            return np.zeros((height, width, 3), dtype=np.uint8)

    assistant = object.__new__(Assistant)
    assistant._bus = get_event_bus()
    assistant._target = _FakeTarget()
    assistant._grabber = _FakeGrabber()

    out = tmp_path / "anchored"
    out.mkdir(parents=True, exist_ok=True)

    node = assistant._capture_start_control(_FakeListener((520, 52)), out, timeout=1.0)
    assert node.handle == 2001
    assert (out / "start_control.json").exists()

    field_map = assistant._build_field_map(2000, node, out)
    assert field_map.has_form
    assert (out / "field_map.json").exists()
    assert (out / "uia_tree.json").exists()


# ---------------------------------------------------------------------------
# MAPPING_RECOVERY + Excel export
# ---------------------------------------------------------------------------


def _source_record_for(rec: dict) -> SourceRecord:
    return SourceRecord(
        pairs={
            "Application Number": rec["key"],
            "Full Name": rec["name"],
            "Gender": rec["gender"],
            "Date Of Birth": rec["dob"],
        },
        ordered_labels=["Application Number", "Full Name", "Gender", "Date Of Birth"],
    )


def _full_field_map(rec: dict) -> UiaFieldMap:
    """A complete field map: every source field binds to a right-form target."""
    from dataclasses import replace

    full = _field_map_for(rec)
    if any(m["source"] == "Application Number" for m in full.mappings):
        return full
    return replace(full, mappings=list(full.mappings) + [
        {"source": "Application Number", "target": "Application Number", "confidence": 0.98},
    ])


def _partial_field_map(rec: dict) -> UiaFieldMap:
    """A field map whose LEFT->RIGHT mappings cover only half the source fields."""
    from dataclasses import replace

    full = _field_map_for(rec)
    return replace(
        full,
        mappings=[m for m in full.mappings if m["source"] in {"Full Name", "Gender"}],
    )


def test_source_coverage_metric_is_source_side() -> None:
    """Coverage is measured on the SOURCE side (the old agent's broken
    ``mapped 21 source fields coverage=46%`` metric), not the form side."""
    rec = {"key": "MPF-100", "name": "KRISHNA", "gender": "Male", "dob": "21 March 1996"}
    record = _source_record_for(rec)
    full = _full_field_map(rec)

    cov, unmapped = AgentLoop._source_coverage(record, full)
    assert cov == 1.0
    assert unmapped == []

    partial = _partial_field_map(rec)
    cov, unmapped = AgentLoop._source_coverage(record, partial)
    assert cov == 0.5
    assert set(unmapped) == {"Application Number", "Date Of Birth"}

    # A record with no valued pairs is trivially 100% (nothing to enter).
    empty = _source_record_for(rec)
    empty.pairs = {}
    cov, unmapped = AgentLoop._source_coverage(empty, full)
    assert cov == 1.0
    assert unmapped == []


def _build_recovery_loop(records: list[dict], refresh) -> AgentLoop:
    get_event_bus().clear()
    target = FieldMapFakeTarget(records)
    controls = RecordingControls()
    plugin = MpfPlugin()
    mapper = SemanticMapper()
    for variant, canonical in plugin._config.get("aliases", {}).items():
        mapper.aliases.learn(variant, canonical)
    executor = ActionExecutor(
        mouse=StubMouse(), keyboard=StubKeyboard(), controls=controls,
        verifier=PassVerifier(), recovery=RecoveryPlanner(),
        verify_after_action=True, max_retries=3, retry_delay=0.0,
    )
    return AgentLoop(
        target=target, source_reader=SourceReader(), mapper=mapper,
        planner=ActionPlanner(verify_after_action=True), executor=executor,
        max_records=1, next_record_timeout=2.0, next_record_poll=0.05,
        field_map=refresh(),
        field_map_refresh=refresh,
        field_driven=True,
    )


def test_mapping_recovery_rebuilds_queue_when_coverage_is_low() -> None:
    """Low source coverage enters MAPPING_RECOVERY and the field map refresh
    brings the queue back to full coverage - never a silent half-fill."""
    rec = {"key": "MPF-100", "name": "KRISHNA", "gender": "Male", "dob": "21 March 1996"}
    calls = {"n": 0}

    def refresh():
        calls["n"] += 1
        # First snapshot has half the mappings; any later snapshot is complete.
        if calls["n"] == 1:
            return _partial_field_map(rec)
        return _full_field_map(rec)

    loop = _build_recovery_loop([rec], refresh)
    record = _source_record_for(rec)
    assert AgentLoop._source_coverage(record, _partial_field_map(rec))[0] < loop._mapping_coverage_threshold

    fresh, queue, attempts = loop._recover_field_driven_mapping(record, 1)
    assert attempts >= 1
    assert loop.state.value == "mapping_recovery"
    cov, unmapped = loop._source_coverage(fresh, loop._field_map)
    assert cov >= loop._mapping_coverage_threshold
    assert unmapped == []
    # The rebuilt queue binds every valued source field.
    bound = {it.label for it in queue.items if it.source_backed}
    assert {"Full Name", "Gender", "Date Of Birth"} <= bound
    # A MAPPING_RECOVERY recovery event was published (state surfaced, not silent).
    assert any(
        ev.data.get("state") == "mapping_recovery"
        for ev in get_event_bus().history(EventType.RECOVERY)
    )


def test_mapping_recovery_is_bounded_and_never_loops() -> None:
    """Even a refresh that never improves coverage stops after the configured
    attempt cap instead of spinning."""
    rec = {"key": "MPF-100", "name": "KRISHNA", "gender": "Male", "dob": "21 March 1996"}
    partial = _partial_field_map(rec)

    loop = _build_recovery_loop([rec], lambda: partial)
    loop._mapping_recovery_max_attempts = 2
    record = _source_record_for(rec)

    fresh, queue, attempts = loop._recover_field_driven_mapping(record, 1)
    assert attempts <= loop._mapping_recovery_max_attempts
    assert queue is not None
    cov, _ = loop._source_coverage(fresh, loop._field_map)
    assert cov < loop._mapping_coverage_threshold  # still low - but no hang


def test_excel_export_appends_one_row_per_record(tmp_path: Path) -> None:
    """The configured workbook gets a header + one row per submitted record with
    the spec's lead columns and every source field."""
    get_event_bus().clear()
    records = [
        {"key": "MPF-100", "name": "KRISHNA", "gender": "Male", "dob": "21 March 1996"},
        {"key": "MPF-200", "name": "RAVI KUMAR", "gender": "Female", "dob": "05 August 1990"},
        # A third record so the final upload advances to a DIFFERENT source key,
        # letting the UIA-first submit verification detect the reset (no VLM).
        {"key": "MPF-300", "name": "SITA DEVI", "gender": "Male", "dob": "14 December 1988"},
    ]
    state = {"idx": 0}

    def refresh():
        current = records[min(state["idx"], len(records) - 1)]
        return _full_field_map(current)

    class SubmitAwareControls(RecordingControls):
        def __init__(self) -> None:
            super().__init__()
            self.uploads = 0

        def click_field(self, bbox, field_id=None):
            if field_id and "uia-btn-3001" in str(field_id):
                state["idx"] = min(state["idx"] + 1, len(records) - 1)
                self.uploads += 1
            return super().click_field(bbox, field_id=field_id)

    controls = SubmitAwareControls()
    plugin = MpfPlugin()
    mapper = SemanticMapper()
    for variant, canonical in plugin._config.get("aliases", {}).items():
        mapper.aliases.learn(variant, canonical)
    executor = ActionExecutor(
        mouse=StubMouse(), keyboard=StubKeyboard(), controls=controls,
        verifier=PassVerifier(), recovery=RecoveryPlanner(),
        verify_after_action=True, max_retries=3, retry_delay=0.0,
    )
    excel = tmp_path / "records.xlsx"
    loop = AgentLoop(
        target=FieldMapFakeTarget(records), source_reader=SourceReader(), mapper=mapper,
        planner=ActionPlanner(verify_after_action=True), executor=executor,
        max_records=2, next_record_timeout=2.0, next_record_poll=0.05,
        field_map=refresh(),
        field_map_refresh=refresh,
        field_driven=True,
        excel_path=str(excel),
        debug_dir=tmp_path,
    )
    summary = loop.run()

    assert summary.completed == 2
    # Every record exported clean coverage (no recovery needed in the full map).
    assert all(r.source_coverage == 1.0 for r in summary.records)
    assert all(r.mapping_recovery_attempts == 0 for r in summary.records)

    from openpyxl import load_workbook

    wb = load_workbook(excel)
    ws = wb["records"]
    header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    assert header[:4] == ["Record Number", "App No", "MBI Code", "Full Name"]
    assert "Gender" in header and "Date Of Birth" in header
    for name in ("Status", "Timestamp", "Verification Status", "Error/Retry Count", "Duration (s)"):
        assert name in header, f"missing meta column {name}"

    def _row(n: int) -> dict:
        return {header[c]: ws.cell(row=n, column=c + 1).value for c in range(len(header))}

    first = _row(2)
    assert first["Record Number"] == "MPF-100"
    assert first["App No"] == "MPF-100"
    assert first["Full Name"] == "KRISHNA"
    assert first["Gender"] == "Male"
    assert first["Date Of Birth"] == "21 March 1996"
    assert first["Status"] == "OK"
    assert first["Verification Status"] == "verified"
    second = _row(3)
    assert second["Record Number"] == "MPF-200"
    assert second["Full Name"] == "RAVI KUMAR"
    # Exactly one row per record (header + 2 data rows).
    assert ws.max_row == 3

