"""The agent workflow loop.

Orchestrates the Observe -> Understand -> Reason -> Plan -> Execute -> Verify
loop for a stream of source records, target-agnostic (desktop window or web
page). It drives every stage explicitly, emits events and state transitions,
and refuses to continue past a record whose actions failed verification.

    while records remain:
        observe    -> target.observe()                       (VLM scene)
        understand -> SourceReader  -> SourceRecord
                     discover_fields -> editable fields
        reason     -> SemanticMapper -> MappingResult
        plan       -> ActionPlanner  -> FillPlan
        execute    -> ActionExecutor  -> verified results
        verify     -> every value-producing action verified (executor)
        next       -> poll until the source record changes, or timeout
"""

from __future__ import annotations

import datetime as _dt
import json
import random
import re
import time
from collections.abc import Callable
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from atlas.act.executor import ActionExecutor
from atlas.act.models import VERIFYABLE_ACTIONS, Action, ActionResult, ActionType
from atlas.core.events import EventType, get_event_bus
from atlas.core.logging import audit_logger, log_screenshot, logger, watchdog_logger
from atlas.core.metrics import Timer
from atlas.core.record_builder import RecordBuilder, RecordBuildResult
from atlas.core.states import AgentState, StateMachine
from atlas.mapping.mapper import MappingResult, SemanticMapper
from atlas.mapping.member_fields import is_member_field
from atlas.mapping.uia_map import UiaFieldMap, pair_source_pairs, PairingDiagnostics
from atlas.mapping.stable_map import LastKnownGoodMapGuard
from atlas.observe.screen_state import build_screen_state
from atlas.observe.source_observer import HARD_FAILURE_CODES
from atlas.observe.uia import ScrollContainer
from atlas.reason.planner import ActionPlanner, FillPlan
from atlas.reason.sections import find_upload_sections
from atlas.target.base import TargetAdapter
from atlas.understanding.fields import discover_fields
from atlas.understanding.source import SourceReader, SourceRecord
from atlas.vision.models import BBox, ElementType, OcrText, SceneDescription, ScreenElement
from atlas.vision.scene import SceneAnalysis
from atlas.workflow.field_engine import (
    DEFAULT_FIELD_RETRIES,
    DEFAULT_FIELD_TIMEOUT,
    DEFAULT_SCROLL_ATTEMPTS,
    DateGroupTarget,
    FieldStatus,
    PendingFieldQueue,
    PerfTracker,
    ProgressGuard,
    ScrollCapabilityCache,
    ScrollProgress,
    TargetNavigator,
    _SUBMIT_OK,
    build_field_actions,
    build_field_queue,
    build_field_queue_from_perception,
    classify_fill_status,
    field_coverage_summary,
    make_scroll_fn,
    source_coverage_from_queue,
)
from atlas.observe.perception import PerceptionStack
from atlas.understanding.target_field import FieldLedgerState, FieldSource, TargetField, control_type_for_uia
from atlas.workflow.audit import AuditStatus, RecordAudit, UploadStatus, build_audit
from atlas.workflow.field_engine import _alias_keys as _engine_alias_keys
from atlas.workflow.ledger import FieldLedger
from atlas.workflow.scroll import PANEL_LEFT, PANEL_RIGHT, DualPanelScroll
from atlas.workflow.scroller import ScrollSession, pick_left_right_containers
from atlas.workflow.viewport import ViewportModel


def _normalized_equals(a: str, b: str) -> bool:
    """Whitespace/separator-insensitive string equality for the pre-filled check."""
    return re.sub(r"\s+", "", a).strip().lower() == re.sub(r"\s+", "", b).strip().lower()


ACTION_STATE = {
    ActionType.TYPE: AgentState.TYPING,
    ActionType.PASTE: AgentState.TYPING,
    ActionType.CLEAR: AgentState.TYPING,
    ActionType.SELECT: AgentState.TYPING,
    ActionType.TOGGLE: AgentState.TYPING,
    ActionType.CHOOSE_DATE: AgentState.TYPING,
    ActionType.TAB: AgentState.TYPING,
    ActionType.PRESS_ENTER: AgentState.TYPING,
    ActionType.PRESS_ESCAPE: AgentState.TYPING,
    ActionType.CLICK: AgentState.CLICKING,
    ActionType.DOUBLE_CLICK: AgentState.CLICKING,
    ActionType.RIGHT_CLICK: AgentState.CLICKING,
    ActionType.HOVER: AgentState.CLICKING,
    ActionType.MOVE_MOUSE: AgentState.CLICKING,
    ActionType.SUBMIT: AgentState.UPLOADING,
    ActionType.SCROLL: AgentState.SCROLLING,
    ActionType.WAIT: AgentState.WAITING,
    ActionType.VERIFY: AgentState.VERIFYING,
    ActionType.CAPTURE: AgentState.ANALYZING,
    ActionType.ANALYZE: AgentState.ANALYZING,
    ActionType.STOP: AgentState.STOPPED,
}

#: Human-readable operation shown on the status panel while each action runs,
#: so the operator always sees the CURRENT operation (never a stale "READING"
#: while the agent is writing / selecting / scrolling).
ACTION_DETAIL = {
    ActionType.TYPE: "WRITING",
    ActionType.PASTE: "WRITING",
    ActionType.CLEAR: "WRITING",
    ActionType.TOGGLE: "WRITING",
    ActionType.CHOOSE_DATE: "WRITING",
    ActionType.TAB: "WRITING",
    ActionType.PRESS_ENTER: "WRITING",
    ActionType.PRESS_ESCAPE: "WRITING",
    ActionType.SELECT: "SELECTING",
    ActionType.CLICK: "CLICKING",
    ActionType.DOUBLE_CLICK: "CLICKING",
    ActionType.RIGHT_CLICK: "CLICKING",
    ActionType.HOVER: "CLICKING",
    ActionType.MOVE_MOUSE: "CLICKING",
    ActionType.SUBMIT: "UPLOADING",
    ActionType.SCROLL: "SCROLLING",
    ActionType.VERIFY: "VERIFYING",
}


@dataclass
class RecordResult:
    """Outcome of processing one source record."""

    index: int
    record: SourceRecord
    mapping: MappingResult
    actions: list[ActionResult] = field(default_factory=list)
    skipped_fields: list[str] = field(default_factory=list)
    success: bool = False
    incomplete_fields: list[str] = field(default_factory=list)
    #: Fields accepted as written with an UNKNOWN verification - tracked and
    #: surfaced (never counted as a verified pass).
    unverified_fields: list[str] = field(default_factory=list)
    duration_ms: float = 0.0
    message: str = ""
    #: Source->form coverage measured on the SOURCE side before filling
    #: (1.0 = every valued source field bound to a form target).
    source_coverage: float = 1.0
    #: Number of MAPPING_RECOVERY attempts performed for this record (0 = none).
    mapping_recovery_attempts: int = 0
    #: Valued source labels that could not be bound to any form target.
    unmapped_source: list[str] = field(default_factory=list)
    #: Final RecordAudit dict (from atlas.workflow.audit), or None when the
    #: record never reached the audit gate.
    audit: dict | None = None

    def to_dict(self) -> dict:
        return {
            "index": self.index,
            "record": self.record.to_dict(),
            "mapping": self.mapping.to_dict(),
            "actions": [a.to_dict() for a in self.actions],
            "skipped_fields": list(self.skipped_fields),
            "success": self.success,
            "incomplete_fields": list(self.incomplete_fields),
            "unverified_fields": list(self.unverified_fields),
            "duration_ms": self.duration_ms,
            "message": self.message,
            "source_coverage": self.source_coverage,
            "mapping_recovery_attempts": self.mapping_recovery_attempts,
            "unmapped_source": list(self.unmapped_source),
            "audit": self.audit,
        }


@dataclass
class WorkflowSummary:
    """Aggregate of a whole workflow run."""

    records: list[RecordResult] = field(default_factory=list)
    started: float = field(default_factory=time.time)
    finished: float = 0.0
    stopped_reason: str = ""

    @property
    def completed(self) -> int:
        return sum(1 for r in self.records if r.success)

    @property
    def failed(self) -> int:
        return sum(1 for r in self.records if not r.success)

    @property
    def unverified(self) -> int:
        """Records containing at least one field written with UNKNOWN verification."""
        return sum(1 for r in self.records if r.unverified_fields)

    @property
    def unverified_fields(self) -> int:
        """Total UNKNOWN-written fields across the run."""
        return sum(len(r.unverified_fields) for r in self.records)

    @property
    def blocked_fields(self) -> list[tuple[str, str, str]]:
        """All value-type/ambiguity-rejected pairings across the run."""
        out: list[tuple[str, str, str]] = []
        for r in self.records:
            out.extend(r.mapping.blocked)
        return out

    @property
    def total_duration(self) -> float:
        return self.finished - self.started if self.finished else 0.0

    @property
    def fields_filled(self) -> int:
        return sum(len(r.actions) for r in self.records)

    def to_dict(self) -> dict:
        return {
            "records": [r.to_dict() for r in self.records],
            "completed": self.completed,
            "failed": self.failed,
            "unverified_records": self.unverified,
            "unverified_fields": self.unverified_fields,
            "blocked_fields": [list(b) for b in self.blocked_fields],
            "total_duration": self.total_duration,
            "stopped_reason": self.stopped_reason,
        }


@dataclass
class RecordContext:
    """Per-record state owned by exactly ONE source record.

    Built fresh for every record and discarded once the form resets to the
    next one, so no old-record field map / queue / status can leak into the
    following record. The loop keeps the record's source signature to detect
    the reset via UIA-first (no VLM).
    """

    index: int
    record: SourceRecord
    field_map: UiaFieldMap | None
    queue: Any | None
    status: str = "active"  # active | filled | submitted | failed
    retry_count: int = 0
    failed_fields: list[str] = field(default_factory=list)

    @property
    def source_signature(self) -> str:
        pairs = sorted(self.record.pairs.items())
        return "|".join(f"{k}={v}" for k, v in pairs)


#: A panel whose structured scroll (pattern/wheel/drag/keyboard/override, all
#: verified) fails this many cycles IN A ROW gets a raw click-and-wheel
#: fallback forced on it regardless of what the structured methods reported -
#: see ``AgentLoop._scroll_one_container``. This is the last line of defence
#: against the panel ever going permanently idle.
_RAW_SCROLL_FAILSAFE_THRESHOLD = 2


class AgentLoop:
    """Runs the observe -> ... -> verify loop until records run out."""

    def __init__(
        self,
        target: TargetAdapter,
        source_reader: SourceReader,
        mapper: SemanticMapper,
        planner: ActionPlanner,
        executor: ActionExecutor,
        memory: Any | None = None,
        verify_after_action: bool = True,
        max_records: int = 0,
        next_record_timeout: float = 120.0,
        next_record_poll: float = 1.5,
        alias_learning: bool = False,
        scene_hook: Callable[[SceneDescription], SceneDescription] | None = None,
        on_record: Callable[[RecordResult], None] | None = None,
        field_map: UiaFieldMap | None = None,
        ocr_callback: Callable[[BBox], list[OcrText]] | None = None,
        debug_dir: str | Path | None = None,
        session_dir: str | Path | None = None,
        state_budget: float | dict[str, float] | None = None,
        record_builder: RecordBuilder | None = None,
        capture_callback: Callable[[Path], bool] | None = None,
        max_scan_rounds: int = 20,
        scan_reveal_fields: bool = False,
        settle_on_start: bool = False,
        scroll_stall_limit: int = 3,
        field_map_refresh: Callable[[], UiaFieldMap | None] | None = None,
        scroll_regions: Callable[[SceneDescription], list[BBox]] | None = None,
        scroll_container_provider: Callable[[], ScrollSession | None] | None = None,
        scroll_min_pixels: int = 250,
        scroll_max_pixels: int = 350,
        scroll_settle: tuple[float, float] = (0.3, 0.5),
        field_driven: bool = False,
        field_driven_scroll: bool = True,
        field_timeout: float = DEFAULT_FIELD_TIMEOUT,
        field_scroll_attempts: int = DEFAULT_SCROLL_ATTEMPTS,
        field_retries: int = DEFAULT_FIELD_RETRIES,
        mapping_coverage_threshold: float = 0.95,
        mapping_recovery_max_attempts: int = 2,
        excel_path: str = "",
        perception_stack: PerceptionStack | None = None,
        second_complete_pass: bool = True,
        single_form: bool = False,
        single_form_upload: bool = False,
        source_observer: Any | None = None,
        source_min_valued_pairs: int = 2,
    ) -> None:
        self._target = target
        self._source_reader = source_reader
        self._mapper = mapper
        self._planner = planner
        self._executor = executor
        self._memory = memory
        self._verify_after_action = verify_after_action
        self._max_records = max_records
        self._next_timeout = next_record_timeout
        self._next_poll = next_record_poll
        self._alias_learning = alias_learning
        self._scene_hook = scene_hook
        self._on_record = on_record
        self._field_map = field_map
        self._ocr_callback = ocr_callback
        self._debug_dir = Path(debug_dir) if debug_dir else None
        self._session_dir = Path(session_dir) if session_dir else (self._debug_dir / "session" if self._debug_dir else None)
        self._record_builder = record_builder or RecordBuilder()
        self._capture_callback = capture_callback
        self._max_scan_rounds = max_scan_rounds
        self._scan_reveal_fields = scan_reveal_fields
        self._settle_on_start = settle_on_start
        self._scroll_stall_limit = max(1, scroll_stall_limit)
        self._field_map_refresh = field_map_refresh
        #: Last-known-good validation for every ``field_map_refresh()`` result
        #: - the concrete fix for the reported field-map explosion (31 left /
        #: 37 -> 79 -> 634 right). See atlas/mapping/stable_map.py. Seeded with
        #: the attach-time map (it is definitionally trusted) so the very
        #: first refresh already has a real baseline to validate against.
        self._map_guard = LastKnownGoodMapGuard()
        self._map_guard.seed(field_map)
        #: Throttle for the expensive full/light field-map rebuild fallback
        #: inside ``_read_source_uia_only`` - only used when the cheap
        #: per-handle value read (``UiaBackend.refresh_source_values``)
        #: cannot refresh any cached source label (PHASE 8/9 fix for the
        #: reported "uia map built" repeating every ~9-10s during ordinary
        #: source-record polling).
        self._last_source_full_refresh: float = 0.0
        self._source_refresh_interval: float = 4.0
        self._scroll_regions = scroll_regions
        self._scroll_min_pixels = max(100, scroll_min_pixels)
        self._scroll_max_pixels = max(self._scroll_min_pixels, scroll_max_pixels)
        self._scroll_settle = scroll_settle
        self._scroll_container_provider = scroll_container_provider
        self._scroll_session: ScrollSession | None = None
        #: Field-driven (performance) path flags. When ``field_driven`` is set
        #: the loop fills from a single ordered UIA field-map queue and scrolls
        #: the RIGHT panel only, instead of the viewport-round reveal pass.
        self._field_driven = field_driven
        self._field_driven_scroll = field_driven_scroll
        self._field_timeout = max(1.0, float(field_timeout))
        self._field_scroll_attempts = max(1, int(field_scroll_attempts))
        self._field_retries = max(0, int(field_retries))
        #: Source-mapping coverage gate. When the share of source fields that
        #: bind to a form target drops below ``_mapping_coverage_threshold`` the
        #: loop enters MAPPING_RECOVERY (re-read source, refresh the UIA field
        #: map) instead of blindly filling whatever happened to map.
        self._mapping_coverage_threshold = max(0.0, min(1.0, float(mapping_coverage_threshold)))
        self._mapping_recovery_max_attempts = max(1, int(mapping_recovery_max_attempts))
        #: Optional Excel workbook for per-record results (see _append_excel_row).
        self._excel_path = (excel_path or "").strip()
        #: Consecutive scroll-method failures per panel (LEFT/RIGHT). When a
        #: panel's structured scroll (pattern/wheel/drag/keyboard/override)
        #: fails this many cycles IN A ROW, `_scroll_one_container` forces a
        #: raw click-and-wheel fallback on the panel's own rect regardless of
        #: what the structured methods think, so the panel can never go
        #: permanently idle. See `_RAW_SCROLL_FAILSAFE_THRESHOLD`.
        self._panel_scroll_failures: dict[str, int] = {}
        self._state_budget = self._normalize_budget(state_budget)
        self._states = StateMachine()
        self._stop = False
        self._pause = False
        self._last_layout = ""
        self._state_entered: dict[AgentState, float] = {}
        self._state_warned: set[AgentState] = set()
        #: Consecutive overrun ticks per state so the level-2 watchdog
        #: escalates instead of logging once and going silent while the loop
        #: keeps spinning inside the same stuck state. Reset on `_set`.
        self._state_overruns: dict[AgentState, int] = {}
        self._last_overrun_log: dict[AgentState, float] = {}
        self._overrun_repeat_log_seconds: float = 30.0
        self._bus = get_event_bus()
        self._cached_analysis: SceneAnalysis | None = None
        self._last_signature = ""
        self._force_rebuild = False
        self._last_field: str | None = None
        self._planner_status = ""
        self._last_exception: str | None = None
        self._no_record_last_reason = ""
        self._expanded_sections: set[str] = set()
        self._scroll_position: int = 0
        self._scroll_blocked_reason: str | None = None
        #: The field-driven queue currently being filled, used to refresh an
        #: action's bbox right before verification (stable-id -> live position).
        self._active_queue: PendingFieldQueue | None = None
        #: RecordContext of the record currently being filled; replaced after
        #: every submit/reset so no stale record state leaks into the next one.
        self._ctx: RecordContext | None = None
        #: How long to wait (UIA-first, no VLM) for the form to reset after a
        #: submit click before falling back to a single VLM confirmation.
        self._submit_reset_timeout: float = 45.0
        #: Expected record count (from --records), used for batch logging.
        self._records_expected: int = 0
        #: Merged perception stack (UIA + CV/OCR fallback) for the second
        #: complete pass and for CV-discovered fields when UIA reports none.
        self._perception_stack = perception_stack
        #: SINGLE-FORM TEST MODE: exactly ONE complete form, then terminate.
        self._single_form = bool(single_form)
        self._single_form_upload = bool(single_form_upload)
        self._single_form_complete = False
        self._terminate_requested = False
        #: Second complete pass: after the fill + completeness passes, re-walk
        #: every source-backed field once more (no-op read-back via the
        #: executor) so a drifted / cascading value is corrected before submit.
        self._second_complete_pass_enabled = second_complete_pass
        #: Per-record field ledger + final audit, set right before the submit
        #: gate. The engine's public ``submit()`` refuses to upload unless the
        #: last audit PASSed (no verified data = no upload).
        self._ledger: FieldLedger | None = None
        self._last_audit: RecordAudit | None = None
        #: Visual LEFT source-panel observer (crop -> OCR -> VLM). When the
        #: UIA/OCR pairing yields no values, the observer reads the record
        #: straight from the source-panel IMAGE so a record is never rejected
        #: just because UIA exposed no label/value rows.
        self._source_observer = source_observer
        if source_observer is not None:
            # Bind the observer's client-rect fallback to the loop's live
            # sandbox rect so a missing UIA left_rect still resolves the LEFT
            # panel by the configured width ratio.
            try:
                source_observer._client_rect_provider = self._client_rect  # type: ignore[attr-defined]
            except Exception:
                pass
        #: Minimum valued pairs for a partial record to count as valid when no
        #: record key is readable (fixes "no record: no valid record detected").
        self._source_min_valued_pairs = max(1, int(source_min_valued_pairs))
        #: Exact reason code of the most recent no-record condition (used to
        #: stop the await loop cleanly instead of spinning forever).
        self._no_record_reason_code: str = ""
        #: Set ONLY when the await loop self-terminates on a hard source
        #: failure, so ``run()`` can distinguish that clean stop from a user
        #: stop (which must keep its "stopped by user" reason).
        self._terminate_reason_code: str = ""
        if self._executor is not None:
            self._executor.set_bbox_refresher(self.refresh_action_bbox)

    # -- lifecycle -----------------------------------------------------------

    @property
    def state(self) -> AgentState:
        return self._states.state

    @property
    def single_form_mode(self) -> bool:
        """True when running the SINGLE-FORM test mode."""
        return self._single_form

    @property
    def single_form_complete(self) -> bool:
        """True once the single form has been fully processed and verified."""
        return self._single_form_complete

    @property
    def terminate_requested(self) -> bool:
        """True when ATLAS should exit cleanly (single-form completion or STOP)."""
        return self._terminate_requested or self._stop

    def request_terminate(self) -> None:
        """Request a clean shutdown at the next safe boundary."""
        self._terminate_requested = True
        self._stop = True

    def stop(self) -> None:
        self._stop = True

    def pause(self) -> None:
        self._pause = True

    def resume(self) -> None:
        self._pause = False

    def run(self) -> WorkflowSummary:
        summary = WorkflowSummary()
        self._states.reset()
        self._bus.publish(EventType.AGENT_STARTED, {"single_form": self._single_form})
        try:
            if not self._target.is_alive():
                raise RuntimeError("target is not attached")
            # Startup: a human looks at the form, waits for it to finish
            # rendering, and only then starts working. Never scroll yet.
            self._wait_until_stable()
            count = 0
            last_key: str | None = None
            self._records_expected = 1 if self._single_form else (self._max_records or 0)
            while not self._stop:
                self._check_state_budget()
                if self._max_records and count >= self._max_records:
                    summary.stopped_reason = f"max_records reached ({count})"
                    break
                if self._pause:
                    time.sleep(0.2)
                    continue
                awaited = self._await_record(last_key)
                if awaited is None:
                    # Reached when the loop was stopped - by the user OR by a
                    # clean hard-failure stop inside ``_await_record``. Only the
                    # latter gets the "no source record readable" reason; a user
                    # stop must keep "stopped by user".
                    if self._terminate_reason_code:
                        summary.stopped_reason = (
                            f"no source record readable [{self._terminate_reason_code}]"
                        )
                    break
                analysis, record = awaited
                try:
                    if self._field_driven:
                        result = self._run_record_field_driven(analysis, record, count + 1)
                    else:
                        result = self._run_record(analysis, record, count + 1)
                except Exception:
                    # Never let a per-record exception swallow the record:
                    # the batch must still report it FAILED (this was the
                    # "BATCH COMPLETE: 0 record(s)" symptom in production).
                    logger.exception("record {} crashed; recording as FAILED", count + 1)
                    result = RecordResult(
                        index=count + 1,
                        record=record,
                        mapping=MappingResult(),
                        success=False,
                        duration_ms=0.0,
                        message="exception during record processing",
                    )
                summary.records.append(result)
                count += 1
                self._append_excel_row(result)
                if self._on_record is not None:
                    try:
                        self._on_record(result)
                    except Exception:
                        logger.exception("on_record callback failed")
                last_key = result.record.record_key
                self._bus.publish(
                    EventType.RECORD_COMPLETED if result.success else EventType.RECORD_FAILED,
                    result.to_dict(),
                )
                logger.info(
                    "record {}/{} {} (App No {}) {} in {:.1f}s",
                    count,
                    self._records_expected or "?",
                    result.record.record_key or "?",
                    result.record.record_key or "?",
                    "OK" if result.success else "FAILED",
                    result.duration_ms / 1000.0,
                )
                if not result.success:
                    logger.warning(
                        "record {}/{} FAILED: {}",
                        count,
                        self._records_expected or "?",
                        result.message or "unknown",
                    )
                # After an upload the left panel changes to the next record.
                # Drop every per-record cache so the next record starts clean
                # (no stale field map / queue / screen model / layout leak).
                self._invalidate_stale_state()
                if self._single_form:
                    # SINGLE-FORM MODE: exactly ONE complete form. After it is
                    # processed (success or failure), never await/load record 2.
                    self._single_form_complete = True
                    self._terminate_requested = True
                    self._stop = True
                    summary.stopped_reason = (
                        "SINGLE FORM COMPLETED — AUTOMATION TERMINATED"
                        if result.success else "SINGLE FORM FAILED"
                    )
                    logger.info(
                        "SINGLE_FORM_MODE: {} - terminating automation cleanly "
                        "(no second record, target application left open)",
                        summary.stopped_reason,
                    )
                    break
        except Exception as exc:
            logger.exception("workflow failed")
            summary.stopped_reason = str(exc)
            self._last_exception = str(exc)
        finally:
            summary.finished = time.time()
            if self._stop:
                summary.stopped_reason = summary.stopped_reason or "stopped by user"
            self._states.transition(AgentState.STOPPED)
            if self._single_form:
                # Single-form terminal state for the dashboard: FINISHED on a
                # fully verified form, ERROR when the form could not be
                # completed (never submit incomplete data).
                sf_ok = bool(summary.records) and all(r.success for r in summary.records)
                self._bus.publish(EventType.STATE_CHANGED, {
                    "state": "finished" if sf_ok else "error",
                    "detail": (
                        "SINGLE FORM COMPLETED — AUTOMATION TERMINATED"
                        if sf_ok else "SINGLE FORM FAILED"
                    ),
                })
            self._bus.publish(EventType.WORKFLOW_COMPLETE, {
                **summary.to_dict(),
                "single_form": self._single_form,
            })
            self._bus.publish(EventType.AGENT_STOPPED, {"reason": summary.stopped_reason})
            logger.info(
                "BATCH COMPLETE: {} record(s) processed ({} OK / {} FAILED) in {:.1f}s - {}",
                len(summary.records),
                summary.completed,
                summary.failed,
                summary.total_duration,
                summary.stopped_reason or "finished",
            )
            self._dump_timeline(summary)
            self._dump_failure(summary)
            self._dump_focus_history()
            self._dump_watchdog()
            self._dump_verification_debug(summary)
            self._dump_run_metrics(summary)
        return summary

    # -- mapping recovery + Excel export ---------------------------------------

    def _enter_mapping_recovery(self, reason: str) -> None:
        """Enter MAPPING_RECOVERY and publish a recovery event.

        Never treated as an error: the loop re-reads the source and refreshes
        the UIA field map (bounded attempts) before deciding anything.
        """
        self._set(AgentState.MAPPING_RECOVERY, "MAPPING_RECOVERY")
        logger.warning("MAPPING_RECOVERY: {}", reason)
        self._bus.publish(EventType.RECOVERY, {"reason": reason, "state": "mapping_recovery"})

    @staticmethod
    def _source_coverage(record: SourceRecord, field_map: UiaFieldMap | None) -> tuple[float, list[str]]:
        """Source-side mapping coverage, member-field driven (FIX #20).

        Of the source labels that carry a value AND resolve to a REQUIRED
        member field, how many bind to a right-form target through the field
        map's LEFT->RIGHT mappings. Returns ``(coverage, unmapped_labels)``.

        Garbage rows (Project Details / Shift Details / timer / buttons) are
        already excluded before this metric runs - see
        ``pair_source_pairs(member_only=True)`` and
        ``SourceObserver._gate_member_pairs`` - and OPTIONAL member fields
        (RAI Code, Mother Tongue, ...) never drag coverage down: a missing
        optional field is not a mapping failure (FIX #9).

        This is the same metric the legacy agent reported as
        ``mapped 21 source fields (coverage=46%)``. Below the configured
        threshold the loop enters MAPPING_RECOVERY instead of blindly filling
        whatever happened to map (never interprets a missing value as "nothing
        to enter").
        """
        from atlas.mapping.member_fields import REQUIRED_MEMBER_FIELDS, resolve_member_field

        pairs = dict(getattr(record, "pairs", {}) or {})
        ordered = list(getattr(record, "ordered_labels", None) or [])
        valued = [label for label in ordered if pairs.get(label)]
        required_valued = [
            label for label in valued if resolve_member_field(label) in REQUIRED_MEMBER_FIELDS
        ]
        if not required_valued:
            return 1.0, []
        mapped = {
            m.get("source")
            for m in (getattr(field_map, "mappings", None) or [])
            if m.get("source") in pairs and pairs.get(m.get("source"))
        }
        bound = [label for label in required_valued if label in mapped]
        unmapped = [label for label in required_valued if label not in mapped]
        return len(bound) / len(required_valued), unmapped

    def _recover_viewport_mapping(
        self,
        record: SourceRecord,
        fields: list[Any],
        index: int,
    ) -> tuple[SourceRecord, MappingResult, int]:
        """Bounded MAPPING_RECOVERY for the viewport path.

        Each attempt re-reads the left source panel UIA-first (fresh pairs),
        refreshes the UIA field map (fresh geometry + mappings) and re-runs the
        semantic mapper. Stops as soon as coverage meets the threshold, or after
        ``_mapping_recovery_max_attempts``. Never loops.
        """
        self._enter_mapping_recovery(
            f"record {index}: source coverage below {self._mapping_coverage_threshold:.0%}"
        )
        fresh_record = record
        mapping = self._mapper.map(record, fields)
        attempts = 0
        for _ in range(self._mapping_recovery_max_attempts):
            attempts += 1
            try:
                self._refresh_field_map_once()
            except Exception as exc:
                logger.debug("mapping recovery field-map refresh failed: {}", exc)
            fresh = self._read_source_uia_only()
            if fresh is not None and len(fresh) > len(fresh_record):
                fresh_record = fresh
            mapping = self._mapper.map(fresh_record, fields)
            logger.info(
                "mapping recovery attempt {}: source coverage {:.0%}",
                attempts,
                mapping.coverage,
            )
            if mapping.coverage >= self._mapping_coverage_threshold:
                break
        return fresh_record, mapping, attempts

    def _recover_field_driven_mapping(
        self,
        record: SourceRecord,
        index: int,
    ) -> tuple[SourceRecord, PendingFieldQueue, int]:
        """Bounded MAPPING_RECOVERY for the field-driven path.

        Each attempt re-reads the source panel UIA-first, refreshes the UIA
        field map and rebuilds the ordered fill queue. Returns the fresh record,
        the rebuilt queue and the number of attempts performed. The queue is
        only returned if a usable form map survived the recovery.
        """
        self._enter_mapping_recovery(
            f"record {index}: source coverage below {self._mapping_coverage_threshold:.0%}"
        )
        fresh_record = record
        queue: PendingFieldQueue | None = None
        attempts = 0
        for _ in range(self._mapping_recovery_max_attempts):
            attempts += 1
            try:
                self._refresh_field_map_once()
            except Exception as exc:
                logger.debug("mapping recovery field-map refresh failed: {}", exc)
            field_map = self._field_map
            if field_map is None or not field_map.has_form:
                break
            fresh = self._read_source_uia_only()
            if fresh is not None and len(fresh) > len(fresh_record):
                fresh_record = fresh
            queue = build_field_queue(field_map, fresh_record)
            coverage, _unmapped = source_coverage_from_queue(fresh_record, queue)
            logger.info(
                "mapping recovery attempt {}: source coverage {:.0%}",
                attempts,
                coverage,
            )
            if coverage >= self._mapping_coverage_threshold:
                break
        if queue is None:
            queue = (
                build_field_queue(self._field_map, fresh_record)
                if self._field_map is not None
                else PendingFieldQueue([])
            )
        return fresh_record, queue, attempts

    #: Fixed leading columns of the Excel export (per the 30-point spec).
    _EXCEL_LEAD = ("Record Number", "App No", "MBI Code", "Full Name")
    #: Trailing metadata columns (source fields slot in between).
    _EXCEL_META = (
        "Status",
        "Timestamp",
        "Verification Status",
        "Error/Retry Count",
        "Duration (s)",
    )

    def _append_excel_row(self, result: RecordResult) -> None:
        """Append one submitted record as a row in the configured workbook.

        Columns: ``Record Number | App No | MBI Code | Full Name | <every source
        field> | Status | Timestamp | Verification Status | Error/Retry Count |
        Duration (s)``. The workbook (and its header) is created on first use;
        new source labels are appended after the fixed lead columns, never
        overwriting existing rows. Failure never aborts the workflow - the row
        is skipped with a warning.
        """
        path = self._excel_path
        if not path:
            return
        try:
            from openpyxl import Workbook, load_workbook
        except ImportError:
            logger.warning("Excel export requested but openpyxl is not installed; row skipped")
            return
        try:
            excel_path = Path(path)
            try:
                excel_path.parent.mkdir(parents=True, exist_ok=True)
            except Exception:
                pass
            exists = excel_path.exists() and excel_path.stat().st_size > 0
            wb = load_workbook(excel_path) if exists else Workbook()
            ws = wb.active
            ws.title = "records"

            pairs = dict(getattr(result.record, "pairs", {}) or {})
            ordered = list(getattr(result.record, "ordered_labels", None) or [])
            lead_extra = _dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            # Locate the canonical lead fields inside the record's own labels.
            def _find_label(*patterns: str) -> str | None:
                lowered = [p.lower() for p in patterns]
                for label in ordered:
                    hay = label.lower()
                    if all(part in hay for part in lowered):
                        return label
                return None

            app_label = _find_label("app", "no") or _find_label("application", "no")
            mbi_label = _find_label("mbi")
            name_label = _find_label("full", "name") or _find_label("name")
            if name_label and name_label.lower().count("name") != 1:
                for label in ordered:
                    low = label.lower()
                    if low in {"full name", "member name", "applicant name", "applicant", "name"}:
                        name_label = label
                        break
            record_key = result.record.record_key or ""
            app_no = pairs.get(app_label, "") if app_label else record_key
            mbi_code = pairs.get(mbi_label, "") if mbi_label else ""
            full_name = pairs.get(name_label, "") if name_label else ""

            # Source fields = everything except the lead columns already taken.
            taken = {app_label, mbi_label, name_label}
            source_header = [label for label in ordered if label not in taken]
            source_values = [pairs.get(label, "") for label in source_header]

            if not exists or ws.max_row < 1 or ws.cell(row=1, column=1).value is None:
                header = list(self._EXCEL_LEAD) + source_header + list(self._EXCEL_META)
                ws.append(header)
            else:
                header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
                known = set(header)
                for label in source_header:
                    if label not in known:
                        ws.cell(row=1, column=ws.max_column + 1).value = label
                        header.append(label)
                        known.add(label)

            lead_values = [record_key, app_no, mbi_code, full_name]
            if not exists or ws.max_row < 2:
                row = lead_values + source_values + [
                    "OK" if result.success else "FAILED",
                    lead_extra,
                    "verified" if not result.unverified_fields
                    else f"{len(result.unverified_fields)} unverified",
                    str(result.mapping_recovery_attempts),
                    f"{result.duration_ms / 1000.0:.1f}",
                ]
                ws.append(row)
            else:
                # Align the row to the (possibly grown) header: fill known lead
                # + source cells by header name, then the trailing meta columns.
                header_lookup = {name: c for c, name in enumerate(header, start=1)}
                row_cells: dict[int, Any] = {}
                for hname in self._EXCEL_LEAD:
                    col = header_lookup.get(hname)
                    if col is not None:
                        row_cells[col] = lead_values[list(self._EXCEL_LEAD).index(hname)]
                for hname, value in zip(source_header, source_values, strict=False):
                    col = header_lookup.get(hname)
                    if col is not None:
                        row_cells[col] = value
                meta_values = [
                    "OK" if result.success else "FAILED",
                    lead_extra,
                    "verified" if not result.unverified_fields
                    else f"{len(result.unverified_fields)} unverified",
                    str(result.mapping_recovery_attempts),
                    f"{result.duration_ms / 1000.0:.1f}",
                ]
                next_row = ws.max_row + 1
                for c, value in row_cells.items():
                    ws.cell(row=next_row, column=c).value = value
                for j, name in enumerate(self._EXCEL_META):
                    col = header_lookup.get(name)
                    if col is not None:
                        ws.cell(row=next_row, column=col).value = meta_values[j]
            wb.save(excel_path)
            logger.info("record {} appended to Excel export: {}", result.index, excel_path)
        except Exception as exc:
            logger.warning("Excel export failed for record {}: {}", result.index, exc)

    # -- record processing ----------------------------------------------------

    def _run_record(self, analysis: SceneAnalysis, record: SourceRecord, index: int) -> RecordResult:
        with Timer() as timer:
            scene = analysis.scene
            self._set(AgentState.SCREEN_MODEL)
            self._set(AgentState.RECORD_EXTRACTION)
            self._bus.publish(EventType.SOURCE_READ, record.to_dict())
            self._bus.publish(EventType.RECORD_STARTED, {"index": index, "record": record.to_dict()})

            # Scroll-locked observation of the current viewport: the whole
            # visible page is read before the first field is touched.
            self._set(AgentState.OBSERVE_VIEWPORT)
            if self._scan_reveal_fields:
                self._lock_scrolling()

            fields = discover_fields(scene)
            if self._scan_reveal_fields:
                # A human scans the current viewport first: below-fold fields are
                # filled later by the reveal pass, so the initial plan must never
                # act on stale geometry it cannot see.
                all_fields = fields
                fields = self._visible_fields(all_fields, self._viewport(analysis))
            else:
                all_fields = fields
            self._bus.publish(
                EventType.FIELD_DISCOVERED, {
                    "count": len(all_fields),
                    "visible": len(fields),
                    "fields": [f.to_dict() for f in fields],
                }
            )

            self._set(AgentState.FIELD_MAPPING)
            mapping = self._mapper.map(record, fields)
            self._bus.publish(EventType.MAPPING, mapping.to_dict())

            recovery_attempts = 0
            if mapping.coverage < self._mapping_coverage_threshold:
                record, mapping, recovery_attempts = self._recover_viewport_mapping(record, fields, index)
                self._bus.publish(EventType.MAPPING, mapping.to_dict())
            result_source_coverage = mapping.coverage
            result_unmapped_source = list(mapping.unmapped_source)
            if result_unmapped_source:
                logger.debug(
                    "record {}: {} source label(s) unmapped: {}",
                    index,
                    len(result_unmapped_source),
                    result_unmapped_source,
                )

            submit_id = self._find_submit(scene)
            self._set(AgentState.PLANNING)
            plan = self._planner.plan_fill(record, mapping, scene, submit_id)
            # When the reveal pass owns the scroll, defer the submit action to
            # the very end so an early click never fires while fields below the
            # fold are still unfilled.
            defer_submit = bool(
                self._scan_reveal_fields
                and submit_id is not None
                and any(self._is_submit_action(a, submit_id) for a in plan.actions)
            )
            if defer_submit:
                plan.actions = [a for a in plan.actions if not self._is_submit_action(a, submit_id)]
            self._bus.publish(EventType.PLAN_CREATED, plan.to_dict())
            self._planner_status = f"{len(plan.actions)} actions planned"

        self._set(AgentState.THINKING)
        key = record.record_key or ""
        self._snapshot("before-fill", index, key)
        self._dump_record_debug(plan, [], index, record)
        results = self._execute_plan(plan, submit_id, index=index, record_key=key)
        # Reveal and fill fields that were below the fold in the first observe
        # (vision targets only see the viewport). Opt-in: the re-observe consumes
        # a fresh screen, which is correct only for live target adapters that
        # return the current screen (idempotent), not sequential test mocks.
        # Bounded; never loops.
        if self._scan_reveal_fields:
            handled_ids = {f.element_id for f in fields}
            handled_labels = {m.source_label for m in mapping.mappings}
            extra = self._scan_fill_revealed(
                record, handled_ids, handled_labels=handled_labels,
                submit_id=submit_id if defer_submit else None, index=index,
                initial_scene=scene,
            )
            if extra:
                results = results + extra
        if not self._all_ok(results):
            self._snapshot("failure", index, key)
        self._snapshot("after-fill", index, key)
        self._dump_record_debug(plan, results, index, record)
        self._bus.publish(
            EventType.SCREEN_STATE,
            build_screen_state(
                scene=scene,
                record=record,
                mapping=mapping,
                results=results,
                window_title=getattr(getattr(self._target, "info", None), "title", "") or "",
                record_index=index,
            ),
        )

        result = RecordResult(
            index=index,
            record=record,
            mapping=mapping,
            actions=results,
            success=self._all_ok(results),
            duration_ms=timer.elapsed * 1000.0,
        )
        result.source_coverage = result_source_coverage
        result.mapping_recovery_attempts = recovery_attempts
        result.unmapped_source = result_unmapped_source
        self._learn_aliases(record, mapping, results)
        result.incomplete_fields = self._unmapped_required(mapping)
        result.skipped_fields = self._skipped_fields(results)
        result.unverified_fields = self._unverified_fields(results)
        if result.skipped_fields or result.incomplete_fields:
            result.message = (
                f"skipped {len(result.skipped_fields)} field(s), "
                f"{len(result.incomplete_fields)} required unmapped"
            )
        if result.unverified_fields:
            suffix = f"; {len(result.unverified_fields)} field(s) written but NOT verified (UNKNOWN)"
            result.message = f"{result.message}{suffix}" if result.message else suffix[2:]
        logger.info(
            "record {} ({}) -> {} in {:.1f}s",
            index,
            record.record_key or "?",
            "OK" if result.success else "FAILED",
            result.duration_ms / 1000.0,
        )
        return result

    # -- field-driven processing (performance path) --------------------------

    def _run_record_field_driven(self, analysis: SceneAnalysis, record: SourceRecord, index: int) -> RecordResult:
        """Fill a record from the ordered UIA field-map queue (performance path).

        Unlike the viewport-round model, the whole form's fields (including
        below-fold ones) come from the UIA field map as ONE ordered queue. The
        loop walks the queue, scrolling the RIGHT panel only to reach below-fold
        targets, and refreshes positions via UIA (no VLM). A single full VLM
        observe is taken after submit to confirm the success indicator / record
        change.
        """
        self._ledger = None
        self._last_audit = None
        with Timer() as timer:
            self._set(AgentState.SCREEN_MODEL)
            self._set(AgentState.RECORD_EXTRACTION)
            self._bus.publish(EventType.SOURCE_READ, record.to_dict())
            self._bus.publish(EventType.RECORD_STARTED, {"index": index, "record": record.to_dict()})

            self._refresh_field_map_once()
            field_map = self._field_map
            if field_map is None or not field_map.has_form:
                self._perception_fallback_note(analysis, "UIA field map insufficient")
                # Phase 6: Try to use perception fields directly instead of viewport fallback
                stack = self._perception_stack
                if stack is not None:
                    handle = None
                    info = getattr(getattr(self, "_target", None), "info", None)
                    if info is not None:
                        try:
                            handle = int(getattr(info, "handle", 0) or 0)
                        except Exception:
                            handle = 0
                    image = None
                    offset = (0, 0)
                    capture = getattr(analysis, "capture", None)
                    if capture is not None:
                        try:
                            image = getattr(capture, "image", None)
                        except Exception:
                            image = None
                        offset = tuple(getattr(capture, "offset", (0, 0)) or (0, 0))
                    try:
                        perception_fields = stack.discover(handle=handle or None, image=image, offset=offset)
                    except Exception as exc:
                        logger.debug("[PERCEPTION] fallback failed: {}", exc)
                        perception_fields = []
                    if perception_fields:
                        logger.info(
                            "field-driven: using {} perception fields for fill queue",
                            len(perception_fields),
                        )
                        queue = build_field_queue_from_perception(perception_fields, record)
                        field_map = None  # no UIA map, using perception queue
                    else:
                        logger.warning("field-driven path needs a UIA field map; falling back to viewport path")
                        return self._run_record(analysis, record, index)
                else:
                    logger.warning("field-driven path needs a UIA field map; falling back to viewport path")
                    return self._run_record(analysis, record, index)

            if field_map is not None:
                queue = build_field_queue(field_map, record)
            if not queue.items:
                logger.warning("field-driven path found no fillable fields; falling back to viewport path")
                return self._run_record(analysis, record, index)

            coverage, unmapped = source_coverage_from_queue(record, queue)
            recovery_attempts = 0
            if coverage < self._mapping_coverage_threshold:
                record, queue, recovery_attempts = self._recover_field_driven_mapping(record, index)
                field_map = self._field_map
                if field_map is None or not field_map.has_form:
                    self._perception_fallback_note(analysis, "MAPPING_RECOVERY lost the form map")
                    logger.warning(
                        "record {}: MAPPING_RECOVERY lost the form map; falling back to viewport path",
                        index,
                    )
                    return self._run_record(analysis, record, index)
                if not queue.items:
                    logger.warning(
                        "record {}: MAPPING_RECOVERY left no fillable fields; falling back to viewport path",
                        index,
                    )
                    return self._run_record(analysis, record, index)
                coverage, unmapped = source_coverage_from_queue(record, queue)
            result_source_coverage = coverage
            result_unmapped_source = list(unmapped)
            logger.info(
                "record {}: source coverage {:.0%} ({} source label(s) unmapped)",
                index,
                coverage,
                len(unmapped),
            )
            if unmapped:
                logger.debug("record {}: unmapped source labels: {}", index, unmapped)

            # A partial source map is not a valid record.  Recovery above is
            # deliberately bounded; if it cannot bind every valued source
            # label with sufficient confidence, stop this record before any
            # submit action rather than silently treating unmatched data as
            # optional.
            if coverage < self._mapping_coverage_threshold:
                reason = (
                    f"source mapping coverage {coverage:.0%} below required "
                    f"{self._mapping_coverage_threshold:.0%}: {', '.join(unmapped)}"
                )
                logger.warning("record {}: submit BLOCKED - {}", index, reason)
                self._write_field_perf(PerfTracker(), queue, index, record, [])
                return RecordResult(
                    index=index,
                    record=record,
                    mapping=MappingResult(),
                    success=False,
                    duration_ms=timer.elapsed * 1000.0,
                    message=reason,
                    source_coverage=coverage,
                    mapping_recovery_attempts=recovery_attempts,
                    unmapped_source=list(unmapped),
                )

            order_ok, bad_at = queue.validate_order()
            if not order_ok:
                logger.warning(
                    "record {}: initial field queue OUT OF reading order at index {} ({} fields)",
                    index,
                    bad_at,
                    len(queue.items),
                )
            self._ctx = RecordContext(
                index=index,
                record=record,
                field_map=field_map,
                queue=queue,
            )

            perf = PerfTracker()
            key = record.record_key or ""
            timings: list[dict] = []
            self._snapshot("before-fill", index, key)
            results = self._fill_from_queue(queue, index, perf, record, timings)

            # Two-pass fill: after the first sequential walk, re-run every
            # source-backed field that failed or was left pending. Cascading
            # dependents (Sub Caste after Caste, Nakshatra after DOB, ...) that
            # were disabled on the first visit are now enabled by their parent
            # and can be filled on the recovery pass.
            blockers = queue.blockers()
            if blockers:
                logger.info(
                    "record {}: completeness pass over {} source-backed field(s)",
                    index,
                    len(blockers),
                )
                self._snapshot("before-retry", index, key)
                for target in list(queue.items):
                    if target.source_backed and (target.failed or target.status is FieldStatus.RETRY_PENDING):
                        queue.mark_status(target, FieldStatus.RETRY_PENDING, "completeness pass retry")
                        target.retries = 0
                extra = self._fill_from_queue(queue, index, perf, record, timings)
                results.extend(extra)
            self._snapshot("after-fill", index, key)

            # Second complete pass: re-walk every source-backed field that the
            # first two passes already verified. The executor's no-op read-back
            # catches a value that drifted (or a cascading dependent that a
            # later field changed) and re-fills it before the final audit.
            if self._second_complete_pass_enabled:
                second = self._second_complete_pass(queue, index, perf, record, timings)
                results.extend(second)

            # Final record audit: rebuild the field ledger from the filled
            # queue and gate the upload through RecordAudit. The engine's
            # submit() guard refuses any upload without a PASS audit.
            self._ledger = self._queue_to_ledger(queue, record)
            self._last_audit = self._audit_record(record, queue, self._ledger, unmapped)
            audit_ok = self._last_audit.allows_submit
            # Publish AUDIT_RESULT event for dashboard
            self._bus.publish(EventType.AUDIT_RESULT, {
                "audit_status": self._last_audit.audit_status.value,
                "upload_status": self._last_audit.upload_status.value,
                "reasons": self._last_audit.reasons,
                "record_index": index,
                "record_key": record.record_key or "",
            })
            if not audit_ok:
                logger.warning(
                    "record {}: submit BLOCKED - final audit {} (upload {})",
                    index,
                    self._last_audit.audit_status.value,
                    self._last_audit.upload_status.value,
                )

            submit_ok: bool | None = None
            blockers = queue.blockers()
            if audit_ok and not blockers and queue.failed == 0 and self._field_map is not None:
                perf.start("submit")
                submit = self._submit_field_driven(record, index)
                perf.stop("submit")
                if submit is not None:
                    results.append(submit)
                    submit_ok = submit.ok
            elif blockers:
                logger.warning(
                    "record {}: submit BLOCKED - {} source-backed field(s) not safely filled",
                    index,
                    len(blockers),
                )
            elif queue.failed:
                logger.warning("record {}: {} field(s) failed; submit skipped", index, queue.failed)

            self._write_field_perf(perf, queue, index, record, timings)

        if self._single_form and not self._single_form_upload:
            # Fill+verify single-form mode: a fully verified form with no
            # blockers/failures is complete even though upload is skipped.
            success = bool(audit_ok and not queue.blockers() and queue.failed == 0)
        else:
            success = bool(not queue.blockers() and queue.failed == 0 and submit_ok is True)
        if self._ctx is not None:
            self._ctx.status = "submitted" if success else "failed"
            self._ctx.failed_fields = [
                it.label for it in queue.items if it.failed
            ]
        result = RecordResult(
            index=index,
            record=record,
            mapping=MappingResult(),
            actions=results,
            success=success,
            duration_ms=timer.elapsed * 1000.0,
        )
        result.source_coverage = result_source_coverage
        result.mapping_recovery_attempts = recovery_attempts
        result.unmapped_source = result_unmapped_source
        result.skipped_fields = self._skipped_fields(results)
        result.unverified_fields = self._unverified_fields(results)
        result.audit = self._last_audit.to_dict() if self._last_audit is not None else None
        parts = []
        if self._single_form and not self._single_form_upload and success:
            parts.append("SINGLE FORM COMPLETED — UPLOAD NOT PERFORMED")
        if self._last_audit is not None and not self._last_audit.allows_submit:
            suffix = "; ".join(self._last_audit.reasons)
            parts.append("final audit BLOCKED" + (f": {suffix}" if suffix else ""))
        blockers = queue.blockers()
        if blockers:
            parts.append(f"{len(blockers)} source-backed field(s) not safely filled")
        if queue.failed:
            parts.append(f"{queue.failed} field(s) failed")
        if result.skipped_fields:
            parts.append(f"{len(result.skipped_fields)} skipped")
        result.message = "; ".join(parts)
        if result.unverified_fields:
            suffix = f"{len(result.unverified_fields)} field(s) written but NOT verified (UNKNOWN)"
            result.message = f"{result.message}; {suffix}" if result.message else suffix
        logger.info(
            "record {} ({}) -> {} in {:.1f}s (field-driven) | statuses={}",
            index,
            record.record_key or "?",
            "OK" if success else "FAILED",
            result.duration_ms / 1000.0,
            queue.status_summary(),
        )
        return result

    def _refresh_field_map_once(self) -> None:
        """Re-query the UIA field map (no VLM) so the queue sees fresh geometry.

        Every result is validated by ``_map_guard`` (PHASE 2/4 fix) before it
        is allowed to replace ``self._field_map``: a candidate that looks like
        a spurious over-walk (right fields exploding while left labels stay
        flat, or internally duplicated fields) is rejected and the previous
        last-known-good map is kept instead. This is the actual fix for the
        observed 31 left / 37 -> 79 -> 634 right field-map corruption - no
        rebuild is trusted purely because it ran.
        """
        if self._field_map_refresh is None:
            return
        try:
            refreshed = self._field_map_refresh()
        except Exception as exc:
            logger.debug("field map refresh failed: {}", exc)
            return
        if refreshed is None:
            return
        accepted_map, report = self._map_guard.evaluate(refreshed)
        if not report.accepted:
            # Rejected: last-known-good map is retained as-is (do not touch
            # self._field_map). Anomaly is already logged by the guard.
            return
        if accepted_map is not None:
            self._field_map = accepted_map

    def _perception_fallback_note(self, analysis: SceneAnalysis, reason: str) -> None:
        """Run the merged perception stack as the UIA-zero fallback channel.

        When the UIA field map exposes no form (UIA insufficient), the
        perception stack (UIA + CV/OCR over the current capture) is consulted
        and the merged channel counts logged/published, so the operator sees
        exactly which channel picked up the form. The fill itself proceeds
        through the vision path (``_run_record``), which fills those fields.
        """
        stack = self._perception_stack
        if stack is None:
            return
        handle = None
        info = getattr(getattr(self, "_target", None), "info", None)
        if info is not None:
            try:
                handle = int(getattr(info, "handle", 0) or 0)
            except Exception:
                handle = 0
        image = None
        offset = (0, 0)
        capture = getattr(analysis, "capture", None)
        if capture is not None:
            try:
                image = getattr(capture, "image", None)
            except Exception:
                image = None
            offset = tuple(getattr(capture, "offset", (0, 0)) or (0, 0))
        try:
            fields = stack.discover(handle=handle or None, image=image, offset=offset)
        except Exception as exc:
            logger.debug("[PERCEPTION] fallback failed: {}", exc)
            return
        by_source: dict[str, int] = {}
        for f in fields:
            by_source[f.source.value] = by_source.get(f.source.value, 0) + 1
        logger.info(
            "[PERCEPTION] fallback active for {}: {} field(s) via {}",
            reason,
            len(fields),
            ", ".join(f"{k}={v}" for k, v in sorted(by_source.items())) or "none",
        )
        self._bus.publish(EventType.FIELD_DISCOVERED, {
            "count": len(fields),
            "visible": len(fields),
            "fields": [f.to_dict() for f in fields],
            "fallback": reason,
        })

    def _field_scroll_session(self) -> ScrollSession | None:
        """Discover the UIA scroll containers once for the field-driven path."""
        if self._scroll_container_provider is None:
            self._scroll_session = None
            return None
        try:
            session = self._scroll_container_provider()
        except Exception as exc:
            logger.debug("scroll container discovery failed: {}", exc)
            session = None
        self._scroll_session = session if (session is not None and session.available) else None
        return self._scroll_session

    def _field_right_container(self, session: ScrollSession | None) -> Any | None:
        """The RIGHT (entry form) scroll container from the discovered session."""
        if session is None or not session.available:
            return None
        field_map = self._field_map
        try:
            chosen = pick_left_right_containers(
                session.containers,
                field_map.left_rect if field_map is not None else None,
                field_map.right_rect if field_map is not None else None,
                self._client_rect(),
            )
        except Exception as exc:
            logger.debug("field-driven container match failed: {}", exc)
            return None
        return chosen.get(PANEL_RIGHT)

    def _field_fill_viewport(
        self,
        session: ScrollSession | None,
        client_rect: tuple[int, int, int, int] | None,
    ) -> tuple[int, int, int, int] | None:
        """The entry form's visible band for fill decisions, or None.

        The RIGHT panel is a UIA scroll container whose ``rect`` is the actual
        on-screen clip region. A fixed status/footer bar often sits just below
        it (the MPF "Record 114 of 114" line) and must never be mistaken for a
        field row - so fill/scroll decisions are made against the container's
        visible band, clipped to the window client rect, instead of the whole
        client area. Falls back to the full client rect when the container is
        unknown (tests / viewport path), preserving the old semantics there.
        """
        if client_rect is None:
            return None
        container = self._field_right_container(session) if session is not None else None
        rect = getattr(container, "rect", None)
        if rect is None:
            return client_rect
        left, top, right, bottom = client_rect
        v_left = max(left, rect.left)
        v_top = max(top, rect.top)
        v_right = min(right, rect.right)
        v_bottom = min(bottom, rect.bottom)
        if v_right <= v_left or v_bottom <= v_top:
            return client_rect
        return (v_left, v_top, v_right, v_bottom)

    def _fill_from_queue(
        self,
        queue: PendingFieldQueue,
        index: int,
        perf: PerfTracker,
        record: SourceRecord,
        timings: list[dict] | None = None,
    ) -> list[ActionResult]:
        """Walk the queue: fill visible targets, scroll RIGHT to reach below-fold ones.

        No field is ever dropped silently: a target with no source value is
        skipped with an explicit NO_SOURCE status (and a FIELD_SKIPPED log),
        and every fill finishes with an explicit status (VERIFIED / FILLED /
        ALREADY_CORRECT / FAILED / RETRY_PENDING).
        """
        results: list[ActionResult] = []
        self._active_queue = queue
        client_rect = self._client_rect()
        guard = ProgressGuard(timeout=self._field_timeout)
        navigator = TargetNavigator(self._scroll_min_pixels, self._scroll_max_pixels)
        cache = ScrollCapabilityCache()
        session = self._field_scroll_session()
        viewport = self._field_fill_viewport(session, client_rect)
        scrolled: dict[str, int] = {}
        total = len(queue.items)
        while not self._stop and queue.next_pending() is not None:
            self._check_state_budget()
            target = queue.next_pending()
            guard.begin()
            label = target.label or target.stable_id
            if not target.source_backed:
                self._mark_skipped(
                    queue, target, FieldStatus.NO_SOURCE, results,
                    f"no source value for {label}",
                )
                done_now = queue.done + queue.failed
                logger.info("[{:>2}/{}] {:<24} SKIP NO_SOURCE", done_now, total, label[:24])
                self._record_timing(timings, target, FieldStatus.NO_SOURCE, 0.0)
                continue
            if navigator.fillable(target, viewport):
                if self._wait_target_enabled(queue, target, guard, navigator, viewport):
                    t0 = time.time()
                    self._bus.publish(
                        EventType.ACTION_STARTED,
                        {"type": "FILL", "field_id": target.stable_id, "label": target.label},
                    )
                    perf.start("fill")
                    ok, action_results = self._fill_target(target, index)
                    perf.stop("fill")
                    results.extend(action_results)
                    elapsed = time.time() - t0
                    self._warn_field_latency(target, elapsed)
                    if ok:
                        status = classify_fill_status(action_results)
                        queue.mark_status(target, status, "")
                        done_now = queue.done + queue.failed
                        logger.info(
                            "[{:>2}/{}] {:<24} OK {:.1f}s [{}]",
                            done_now, total, label[:24], elapsed, status.value,
                        )
                        self._record_timing(timings, target, status, elapsed)
                        continue
                    target.retries += 1
                    if target.retries > self._field_retries:
                        self._mark_failed(queue, target, results, "fill failed")
                        self._record_timing(timings, target, FieldStatus.FAILED, elapsed)
                        continue
                    queue.mark_status(target, FieldStatus.RETRY_PENDING, "fill failed once - retrying")
                    self._record_timing(timings, target, FieldStatus.RETRY_PENDING, elapsed)
                    self._refresh_and_merge(queue, record)
                    continue
                self._mark_failed(queue, target, results, "dependent field never enabled")
                self._record_timing(timings, target, FieldStatus.FAILED, 0.0)
                continue
            # Below the fold: scroll the RIGHT panel toward the target.
            if not self._field_driven_scroll:
                self._mark_failed(queue, target, results, "target below fold and scroll disabled")
                self._record_timing(timings, target, FieldStatus.FAILED, 0.0)
                continue
            tries = scrolled.get(target.stable_id, 0)
            if tries >= self._field_scroll_attempts:
                self._mark_failed(queue, target, results, "target below fold after repeated scrolling")
                self._record_timing(timings, target, FieldStatus.FAILED, 0.0)
                continue
            perf.start("scroll")
            moved = self._scroll_to_target(queue, target, viewport, session, cache, navigator, guard)
            perf.stop("scroll")
            scrolled[target.stable_id] = tries + 1
            if not moved and not navigator.fillable(target, viewport):
                self._mark_failed(queue, target, results, "could not scroll target into view")
                self._record_timing(timings, target, FieldStatus.FAILED, 0.0)
                continue
        return results

    @staticmethod
    def _record_timing(timings: list[dict] | None, target: Any, status: FieldStatus, elapsed: float) -> None:
        """Append one per-field timing row to the report collector."""
        if timings is None:
            return
        timings.append({
            "stable_id": target.stable_id,
            "label": target.label or target.stable_id,
            "status": status.value,
            "elapsed": round(elapsed, 3),
        })

    @staticmethod
    def _warn_field_latency(target: Any, elapsed: float) -> None:
        """Surface a slow field so a drag never hides behind the progress logs."""
        label = target.label or target.stable_id
        if elapsed >= 5.0:
            logger.warning("VERY_SLOW_FIELD {:<24} {:.1f}s", label, elapsed)
        elif elapsed >= 3.0:
            logger.warning("SLOW_FIELD {:<24} {:.1f}s", label, elapsed)

    def _refresh_and_merge(self, queue: PendingFieldQueue, record: SourceRecord) -> None:
        """Refresh UIA positions AND fold newly discovered fields into the queue.

        The queue was built from the first snapshot; fields that only appear
        later (lazy render, dynamic sections) must join it or they are skipped
        forever. ``merge_fields`` appends them in reading order without
        disturbing the already-built deterministic order.
        """
        self._refresh_field_map_once()
        if self._field_map is None:
            return
        queue.refresh_positions(self._field_map.right_fields)
        added = queue.merge_fields(self._field_map.right_fields)
        if added:
            logger.info("field queue grew by {} field(s) after refresh", added)

    def _wait_target_enabled(
        self,
        queue: PendingFieldQueue,
        target: Any,
        guard: ProgressGuard,
        navigator: TargetNavigator,
        viewport: tuple[int, int, int, int] | None,
    ) -> bool:
        """Wait for a cascading/dependent field to become enabled (bounded).

        A dependent dropdown (District after State, Taluk after District,
        Caste after Religion, Sub Caste after Caste, Nakshatra after DOB, ...)
        starts DISABLED until its parent is filled. Refreshing only the UIA
        positions (no VLM) lets the live ``enabled`` state propagate into the
        queue's nodes, with adaptive polling that backs off as the wait drags
        on. Bounded by the per-field progress guard so a genuinely dead field
        is failed quickly instead of stalling the record.
        """
        if target.enabled:
            return True
        logger.info(
            "field {} is disabled - waiting for dependent parent to enable it",
            target.stable_id,
        )
        intervals = (0.15, 0.25, 0.4, 0.6, 0.9)
        idx = 0
        # The dependent-dropdown budget (<5s) is much tighter than the general
        # per-field guard, so a dead disabled field fails fast instead of
        # stalling the record.
        deadline = time.time() + min(5.0, self._field_timeout)
        while not self._stop and not guard.expired and time.time() < deadline:
            # NOTE: no early-return when the target scrolls away mid-wait. The
            # fill loop never scrolls during this wait, so a target that
            # entered fillable stays fillable; returning True on a non-fillable
            # target used to let a STILL-DISABLED field be marked done with no
            # actions (a silent skip of the very dependent combos this wait
            # exists to protect).
            time.sleep(intervals[idx])
            idx = min(idx + 1, len(intervals) - 1)
            self._refresh_field_map_once()
            if self._field_map is not None:
                queue.refresh_positions(self._field_map.right_fields)
            if target.enabled:
                return True
        return target.enabled

    def _scroll_to_target(
        self,
        queue: PendingFieldQueue,
        target: Any,
        viewport: tuple[int, int, int, int] | None,
        session: ScrollSession | None,
        cache: ScrollCapabilityCache,
        navigator: TargetNavigator,
        guard: ProgressGuard,
    ) -> bool:
        """Scroll the RIGHT panel toward a below-fold target using the cached method.

        Verifies progress via container percent / target-y change (no full VLM
        re-observe). Returns True when the target became fillable or moved.
        """
        container = self._field_right_container(session)
        if container is None:
            return self._wheel_fallback_to_target(queue, target, viewport, navigator, guard)
        amount = navigator.scroll_amount_for(target, viewport)
        before = ScrollProgress.capture(container, target)
        self._set(AgentState.SCROLLING, "SCROLLING FORM (FIELD-DRIVEN)")
        scroller = getattr(session, "scroller", None)
        backend = getattr(scroller, "_backend", None)
        dom_available = getattr(scroller, "_dom", None) is not None
        info = getattr(self._target, "info", None)
        handle = getattr(info, "handle", None)
        method = cache.method_for(container, dom_available=dom_available)

        for _ in range(self._field_scroll_attempts):
            if guard.expired:
                break
            fn = make_scroll_fn(session, method, handle=handle, backend=backend)
            moved = bool(fn(container, amount)) if fn is not None else False
            if not moved:
                outcome = None
                if scroller is not None:
                    try:
                        outcome = scroller.scroll_down(container, amount, verify=lambda: False)
                    except Exception as exc:
                        logger.debug("escalation scroll failed: {}", exc)
                method = outcome.method if outcome is not None else "none"
                moved = bool(outcome is not None and outcome.method != "none")
            if method != "none":
                cache.remember(container, method)
            if moved and backend is not None and getattr(container, "has_scroll_pattern", False):
                try:
                    backend.container_state(container, handle)
                except Exception:
                    pass
            time.sleep(random.uniform(*self._scroll_settle))
            self._refresh_field_map_once()
            if self._field_map is not None:
                queue.refresh_positions(self._field_map.right_fields)
                queue.merge_fields(self._field_map.right_fields)
            if navigator.fillable(target, viewport) or before.moved(container, target):
                return True
            before = ScrollProgress.capture(container, target)
        return navigator.fillable(target, viewport)

    def _wheel_fallback_to_target(
        self,
        queue: PendingFieldQueue,
        target: Any,
        viewport: tuple[int, int, int, int] | None,
        navigator: TargetNavigator,
        guard: ProgressGuard,
    ) -> bool:
        """No UIA scroll container: click-focus + wheel the RIGHT panel rect."""
        field_map = self._field_map
        rect = field_map.right_rect if field_map is not None else None
        if rect is None:
            return False
        amount = navigator.scroll_amount_for(target, viewport)
        self._set(AgentState.SCROLLING, "SCROLLING FORM (FIELD-DRIVEN)")
        for _ in range(self._field_scroll_attempts):
            if guard.expired:
                break
            self._scroll_region(PANEL_RIGHT, rect, amount, reason="scroll right panel toward field")
            time.sleep(random.uniform(*self._scroll_settle))
            self._refresh_field_map_once()
            if self._field_map is not None:
                queue.refresh_positions(self._field_map.right_fields)
                queue.merge_fields(self._field_map.right_fields)
            if navigator.fillable(target, viewport):
                return True
        return navigator.fillable(target, viewport)

    def refresh_action_bbox(self, field_id: str | None) -> BBox | None:
        """Live bbox for an action by stable-id match against the active queue.

        Called by the executor just before verification so a read never uses a
        bbox made stale by a scroll or window resize since the write. Returns
        None outside the field-driven path (the viewport path re-observes
        instead), so no stale geometry is ever substituted there.
        """
        if field_id is None or self._active_queue is None:
            return None
        try:
            return self._active_queue.bbox_for_id(field_id)
        except Exception:
            return None

    def _fill_target(self, target: Any, index: int) -> tuple[bool, list[ActionResult]]:
        """Execute the fill actions for one target; returns (ok, action_results).

        Pre-filled skip (FIX #20): if the control ALREADY holds the source
        value, the write is a genuine no-op and ZERO mouse/keyboard actions are
        performed - the UIA value is authoritative and cheaper than a VLM
        no-op read. Reported as ALREADY_CORRECT (verified pass, satisfies the
        submit gate).
        """
        current = getattr(target, "current_value", None)
        wanted = getattr(target, "value", None)
        if (
            current is not None
            and wanted is not None
            and str(current or "").strip() and str(wanted or "").strip()
            and _normalized_equals(str(current), str(wanted))
        ):
            label = target.label or target.stable_id
            logger.info("[SKIP] {} already populated (no-op, zero input)", label)
            skip_result = ActionResult(
                action=Action(
                    type=ActionType.STOP,
                    field_id=target.stable_id,
                    reason=f"{label} already populated",
                ),
                success=True,
                verified=True,
                verification_status="ALREADY_CORRECT",
                message=f"{label} already populated - zero input performed",
            )
            return True, [skip_result]
        actions = build_field_actions(target)
        results: list[ActionResult] = []
        for action in actions:
            if self._stop:
                break
            self._last_field = action.field_id or action.reason
            self._set(
                ACTION_STATE.get(action.type, AgentState.THINKING),
                ACTION_DETAIL.get(action.type),
            )
            self._bus.publish(EventType.ACTION_STARTED, action.to_dict())
            result = self._executor.execute(action)
            results.append(result)
            if not result.ok:
                logger.warning("field fill action failed: {} ({})", action.reason, result.message)
                return False, results
        # Whole-group verification for date triplets: after the per-part fills,
        # one combined read over Day+Month+Year confirms the date as a whole.
        # The date-aware verifier lets any spelling match (source ISO "1996-02-02"
        # vs. the triplet's "02 02 1996"), and the union bbox is refreshed live.
        if isinstance(target, DateGroupTarget) and getattr(target, "date_value", None):
            group_bbox = target.bbox
            if group_bbox is not None:
                verify = Action(
                    type=ActionType.VERIFY,
                    field_id=target.stable_id,
                    bbox=group_bbox,
                    value=target.date_value,
                    expected=target.date_value,
                    reason="verify DOB date group",
                )
                if not self._stop:
                    self._last_field = verify.field_id
                    self._set(AgentState.VERIFYING, "VERIFYING")
                    self._bus.publish(EventType.ACTION_STARTED, verify.to_dict())
                    result = self._executor.execute(verify)
                    results.append(result)
                    if not result.ok:
                        logger.warning("field fill action failed: {} ({})", verify.reason, result.message)
                        return False, results
        return all(r.ok for r in results), results

    def _mark_failed(self, queue: PendingFieldQueue, target: Any, results: list[ActionResult], reason: str) -> None:
        queue.mark_failed(target, reason)
        logger.warning("field {} failed: {}", getattr(target, "stable_id", target), reason)
        results.append(ActionResult(
            action=Action(type=ActionType.STOP, reason=reason),
            success=False,
            verified=False,
            message=reason,
        ))

    def _mark_skipped(
        self,
        queue: PendingFieldQueue,
        target: Any,
        status: FieldStatus,
        results: list[ActionResult],
        reason: str,
    ) -> None:
        """Record a deliberate skip with an explicit status and reason.

        Every skip is surfaced: logged (FIELD_SKIPPED), stored on the queue's
        ``skipped_items`` for the completeness report, and appended to the
        record's action results so the summary can account for it.
        """
        queue.mark_skipped(target, status, reason)
        logger.info(
            "FIELD_SKIPPED {}: {} [{}]",
            getattr(target, "stable_id", target),
            reason,
            status.value,
        )
        results.append(ActionResult(
            action=Action(type=ActionType.STOP, reason=reason, field_id=getattr(target, "stable_id", None)),
            success=True,
            verified=False,
            verification_status=status.value,
            message=reason,
        ))

    def _submit_field_driven(self, record: SourceRecord, index: int) -> ActionResult | None:
        """Click the upload button once, then verify with a single VLM observe.

        The engine-level ``submit()`` guard runs first: no PASS final audit, no
        upload (the second upload protection, alongside the UI button state).

        In SINGLE-FORM mode with upload disabled this method NEVER clicks the
        button - it returns immediately and publishes a blocked event, leaving
        the completed form on screen for inspection.
        """
        if self._single_form and not self._single_form_upload:
            logger.info("UPLOAD BLOCKED — SINGLE_FORM_MODE; form left filled and verified")
            self._bus.publish(EventType.UPLOADING, {
                "blocked": True,
                "reason": "SINGLE_FORM_MODE - upload not performed",
                "record": index,
            })
            return None
        if not self.submit():
            audit_logger.warning(
                "submit BLOCKED (engine guard): final audit {} / upload {}",
                self._last_audit.audit_status.value if self._last_audit else "N/A",
                self._last_audit.upload_status.value if self._last_audit else "N/A",
            )
            self._bus.publish(EventType.UPLOADING, {
                "blocked": True,
                "reason": "final audit did not pass",
                "audit": self._last_audit.to_dict() if self._last_audit else None,
            })
            return None
        field_map = self._field_map
        if field_map is None or field_map.upload_button is None or field_map.upload_button.rect is None:
            logger.warning("no upload button in field map; submit skipped")
            return None
        button = field_map.upload_button
        action = Action(
            type=ActionType.CLICK,
            field_id=f"uia-btn-{button.handle or 'upload'}",
            bbox=button.rect,
            confidence=1.0,
            expected="clicked submit",
            reason="click submit button after filling the form",
        )
        self._bus.publish(EventType.UPLOADING, action.to_dict())
        self._set(AgentState.READY_TO_SUBMIT, "READY_TO_SUBMIT")
        self._last_field = action.field_id or action.reason
        result = self._executor.execute(action)
        self._set(AgentState.SUBMITTING, "SUBMITTING")
        if result.ok:
            self._bus.publish(EventType.UPLOAD_COMPLETED, result.to_dict())
            self._snapshot("after-upload", index, record.record_key or "")
        result.success = result.ok and self._verify_submit(record)
        return result

    def _verify_submit(self, record: SourceRecord) -> bool:
        """Verify the submit by waiting for the form to reset to the next record.

        UIA-first: refresh the left source panel and watch for a source-signature
        change (new App No / cleared form) each poll - no VLM, no OCR. Only after
        ``_submit_reset_timeout`` does it take ONE full VLM re-observe and fall
        back to the legacy success/error-token check. Returns True only on
        evidence of a successful submit + reset; never blocks forever.
        """
        old_key = record.record_key or ""
        old_sig = sorted(record.pairs.items())
        self._set(AgentState.SUBMIT_VERIFICATION, "SUBMIT_VERIFICATION")
        deadline = time.time() + self._submit_reset_timeout
        last_wait_log = 0.0
        while not self._stop and time.time() < deadline:
            self._check_state_budget()
            self._set(AgentState.WAITING_FOR_RESET, "WAITING_FOR_RESET")
            next_record = self._read_source_uia_only()
            if next_record is not None:
                new_key = next_record.record_key or ""
                new_sig = sorted(next_record.pairs.items())
                if (new_key and new_key != old_key) or new_sig != old_sig:
                    self._publish_reset(old_key, new_key)
                    self._invalidate_stale_state()
                    return True
                # Same record still showing: the form has not reset yet.
            now = time.time()
            if now - last_wait_log >= 10.0:
                last_wait_log = now
                logger.info(
                    "submit: waiting for form reset ({}s remaining)...",
                    int(deadline - now),
                )
            if not self._target.is_alive():
                logger.warning("target window disappeared during submit verification")
                return False
            time.sleep(self._next_poll)

        # Timeout: single VLM fallback (legacy success/error-token check).
        logger.info("submit: reset not detected via UIA; one VLM confirmation observe")
        self._force_rebuild = True
        try:
            analysis, _ = self._observe()
        except Exception as exc:
            logger.debug("post-submit observe failed: {}", exc)
            analysis = None
        if analysis is None:
            return True
        scene = analysis.scene
        try:
            next_record = self._extract_record(scene)
        except Exception as exc:
            logger.debug("post-submit record read failed: {}", exc)
            next_record = None
        if next_record is not None and next_record.record_key and next_record.record_key != record.record_key:
            self._publish_reset(old_key, next_record.record_key)
            self._invalidate_stale_state()
            return True
        text = " ".join(
            f"{e.label or ''} {e.name or ''}".lower()
            for e in scene.elements
        )
        if any(token in text for token in (
            "success", "submitted", "saved successfully", "record created",
            "upload successful", "record saved", "data saved",
        )):
            return True
        return not any(token in text for token in ("error", "failed", "validation error", "invalid", "cannot be blank"))

    # -- final audit + engine submit() guard ----------------------------------

    def audit(self) -> RecordAudit | None:
        """The most recent per-record final audit (None until a record audits)."""
        return self._last_audit

    def ledger(self) -> FieldLedger | None:
        """The most recent per-record field ledger (None until a record fills)."""
        return self._ledger

    def allows_submit(self) -> bool:
        """Engine-level upload gate: True only after a PASS final audit."""
        return self._last_audit is not None and self._last_audit.allows_submit

    def submit(self) -> bool:
        """Engine-level upload guard. Refuses to submit unless the final audit
        for the current record PASSed. No verified data = no upload.

        In SINGLE-FORM mode with upload disabled this guard is absolute: no
        audit result can permit an upload, so "Upload Details" is never
        clicked through any path (plan action, field-driven submit, fallback).
        """
        if self._single_form and not self._single_form_upload:
            audit_logger.warning("UPLOAD BLOCKED — SINGLE_FORM_MODE (upload not permitted)")
            return False
        if not self.allows_submit():
            audit_logger.warning(
                "submit() rejected: {}",
                "; ".join(self._last_audit.reasons) if self._last_audit else "no audit for current record",
            )
            return False
        return True

    @staticmethod
    def _field_state_for_status(status: FieldStatus) -> FieldLedgerState:
        """Map the queue's terminal FieldStatus onto the ledger lifecycle."""
        if status is FieldStatus.VERIFIED or status is FieldStatus.ALREADY_CORRECT:
            return FieldLedgerState.VERIFIED
        if status is FieldStatus.FAILED:
            return FieldLedgerState.FAILED
        if status is FieldStatus.UNMAPPED:
            return FieldLedgerState.UNMAPPED
        if status is FieldStatus.NOT_APPLICABLE or status is FieldStatus.NO_SOURCE:
            return FieldLedgerState.SKIPPED
        if status is FieldStatus.FILLED or status is FieldStatus.RETRY_PENDING:
            return FieldLedgerState.ENTERED
        return FieldLedgerState.DISCOVERED

    @staticmethod
    def _queue_sub_targets(item: Any) -> list[Any]:
        """The field targets behind one queue item (DateGroupTarget flattens)."""
        return list(item.targets) if isinstance(item, DateGroupTarget) else [item]

    def _second_complete_pass(
        self,
        queue: PendingFieldQueue,
        index: int,
        perf: PerfTracker,
        record: SourceRecord,
        timings: list[dict] | None = None,
    ) -> list[ActionResult]:
        """Re-verify every source-backed field already marked VERIFIED.

        First a read-only VERIFY over each field (no write): a value that
        drifted - or a cascading dependent a later field changed - is detected
        and re-filled; a correct value stays untouched. Bounded to one pass.
        """
        targets = [
            t for item in queue.items
            for t in self._queue_sub_targets(item)
            if t.source_backed and t.status in _SUBMIT_OK
        ]
        if not targets:
            return []
        logger.info(
            "record {}: second complete pass - re-verifying {} field(s)",
            index,
            len(targets),
        )
        self._active_queue = queue
        stale: list[Any] = []
        for t in targets:
            if self._stop:
                break
            bbox = t.bbox
            if bbox is None:
                stale.append(t)
                continue
            action = Action(
                type=ActionType.VERIFY,
                field_id=t.stable_id,
                bbox=bbox,
                value=t.value,
                expected=t.value,
                reason=f"second complete pass: {t.label or t.stable_id}",
            )
            try:
                vresult = self._executor.execute(action)
            except Exception as exc:
                logger.debug("second complete pass verify failed for {}: {}", t.label, exc)
                vresult = None
            if vresult is None or not vresult.verified:
                stale.append(t)
        if not stale:
            logger.info(
                "record {}: second complete pass - all {} field(s) verified",
                index,
                len(targets),
            )
            return []
        logger.warning(
            "record {}: second complete pass - {} field(s) drifted; re-filling",
            index,
            len(stale),
        )
        for item in queue.items:
            for t in self._queue_sub_targets(item):
                if t in stale:
                    queue.mark_status(t, FieldStatus.RETRY_PENDING, "second complete pass refill")
                    t.retries = 0
        return self._fill_from_queue(queue, index, perf, record, timings)

    def _queue_to_ledger(self, queue: PendingFieldQueue, record: SourceRecord) -> FieldLedger:
        """Rebuild the per-record FieldLedger from the filled queue.

        The queue is the single source of truth after execution, so the ledger
        registers every target with its terminal field state. Source->target
        bindings are reconstructed with the same alias matching the coverage
        metric uses, keeping the audit consistent with the existing gates.
        """
        ledger = FieldLedger()
        for item in queue.items:
            for t in self._queue_sub_targets(item):
                field = TargetField(
                    id=t.stable_id,
                    label=t.label,
                    normalized_label=self._normalize_ledger_label(t.label),
                    section=t.section,
                    control_type=control_type_for_uia(t.control_type),
                    bounds=t.bbox,
                    value=t.value,
                    options=t.options,
                    enabled=t.enabled,
                    source=FieldSource.UIA,
                    state=self._field_state_for_status(t.status),
                    ref=t,
                )
                ledger.register(field)
        for source_label, field_id in self._build_source_map(record, queue).items():
            ledger.register_mapping(source_label, field_id)
        return ledger

    @staticmethod
    def _normalize_ledger_label(label: str) -> str:
        return re.sub(r"[:：\s]+$", "", (label or "")).strip().lower()

    def _build_source_map(self, record: SourceRecord, queue: PendingFieldQueue) -> dict[str, str]:
        """Normalized source label -> stable target id for the audit.

        Mirrors ``source_coverage_from_queue`` alias matching so the audit's
        unmapped/missing sets never contradict the coverage gate the queue
        already enforced.
        """
        pairs = dict(getattr(record, "pairs", {}) or {})
        ordered = list(getattr(record, "ordered_labels", None) or [])
        valued = [label for label in ordered if str(pairs.get(label) or "").strip()]
        source_map: dict[str, str] = {}
        for item in queue.items:
            if not getattr(item, "source_backed", False):
                continue
            subs = self._queue_sub_targets(item)
            labels = [getattr(t, "label", "") or "" for t in subs]
            item_keys: set[str] = set()
            for label in labels:
                item_keys.update(_engine_alias_keys(label))
            for source_label in valued:
                if _engine_alias_keys(source_label) & item_keys:
                    source_map.setdefault(source_label.strip().lower(), subs[0].stable_id)
        return source_map

    def _audit_record(
        self,
        record: SourceRecord,
        queue: PendingFieldQueue,
        ledger: FieldLedger,
        unmapped: list[str] | None = None,
    ) -> RecordAudit:
        """Build the final RecordAudit from the ledger + source record.

        ``source_labels`` are the valued source labels (the audit's job is to
        guarantee every one of them reached a VERIFIED target field).
        """
        pairs = dict(getattr(record, "pairs", {}) or {})
        ordered = list(getattr(record, "ordered_labels", None) or [])
        source_labels = [label for label in ordered if str(pairs.get(label) or "").strip()]
        return build_audit(
            source_labels=source_labels,
            fields=ledger.fields,
            source_map=ledger.source_map,
        )

    def _publish_reset(self, old_key: str, new_key: str) -> None:
        """Log + publish a form-reset (old record gone, source panel changed)."""
        self._set(AgentState.RESET_DETECTED, "RESET_DETECTED")
        logger.info("RESET DETECTED: {} -> {}", old_key or "?", new_key or "?")
        self._bus.publish(EventType.RECOVERY, {
            "reason": f"record reset {old_key or '?'} -> {new_key or '?'}",
            "state": "reset_detected",
        })

    def _log_reset_transition(self, previous_key: str | None, record: SourceRecord) -> None:
        """Log the record-key transition observed by the await loop."""
        key = record.record_key or ""
        if previous_key and key and previous_key != key:
            logger.info("NEXT RECORD: {} -> {}", previous_key, key)

    def _invalidate_stale_state(self) -> None:
        """Drop every per-record cache so the next record starts clean.

        Old field-map geometry, the fill queue, the cached screen model and the
        layout fingerprint all belong to the record that just left; keeping any
        of them would leak stale data into the next record.
        """
        self._ctx = None
        self._active_queue = None
        self._ledger = None
        self._last_audit = None
        self._last_layout = ""
        self._cached_analysis = None
        self._force_rebuild = True
        self._no_record_last_reason = ""

    def _write_field_perf(
        self, perf: PerfTracker, queue: PendingFieldQueue, index: int, record: SourceRecord,
        timings: list[dict] | None = None,
    ) -> None:
        coverage = field_coverage_summary(queue)
        logger.info(
            "record {}: targets={} mapped={} ({:.0%}) statuses={} skipped={} failed={}",
            index,
            coverage["total_targets"],
            coverage["mapped_targets"],
            coverage["mapped_pct"],
            queue.status_summary(),
            len(queue.skipped_items),
            queue.failed,
        )
        if self._debug_dir is None:
            return
        self._write_debug("field_driven_perf.json", {
            "record_index": index,
            "key": record.record_key,
            "phases": perf.to_dict(),
            "coverage": coverage,
            "timings": timings or [],
            "queue": {
                "total": len(queue.items),
                "done": queue.done,
                "failed": queue.failed,
                "remaining": queue.remaining,
                "skipped": len(queue.skipped_items),
                "statuses": queue.status_summary(),
                "blockers": [
                    {
                        "label": it.label,
                        "status": it.status.value,
                        "reason": it.status_reason,
                    }
                    for it in queue.blockers()
                ],
            },
        })

    def _execute_plan(
        self, plan: FillPlan, submit_element_id: str | None = None,
        index: int = 0, record_key: str = "",
    ) -> list[ActionResult]:
        results: list[ActionResult] = []
        for action in plan.actions:
            if self._stop:
                break
            self._check_state_budget()
            is_upload = action.type == ActionType.SUBMIT or (
                action.type == ActionType.CLICK and action.field_id == submit_element_id
            )
            if is_upload and self._single_form and not self._single_form_upload:
                # Hard safety guard (spec 16): single-form fill+verify mode must
                # never trigger the upload button through ANY path.
                logger.info("UPLOAD BLOCKED — SINGLE_FORM_MODE; skipping upload action")
                self._bus.publish(EventType.UPLOADING, {
                    "blocked": True,
                    "reason": "SINGLE_FORM_MODE - upload not performed",
                    "record": index,
                })
                break
            if is_upload:
                # Hard completion gate (legacy-path safety net): never submit
                # if any value-bearing action taken so far in this record
                # failed outright or was only accepted with an UNKNOWN
                # verification (written but not confirmed). This mirrors the
                # field-driven queue.blockers() gate for the plan-based path,
                # which otherwise has no submit gate at all.
                unresolved = [
                    r for r in results
                    if r.action.type in VERIFYABLE_ACTIONS
                    and r.action.value is not None
                    and not r.verified
                ]
                if unresolved:
                    names = ", ".join(
                        (r.action.reason or r.action.field_id or "?") for r in unresolved[:10]
                    )
                    logger.warning(
                        "record {}: submit BLOCKED (legacy path) - {} field(s) not verified: {}",
                        index, len(unresolved), names,
                    )
                    break
                self._bus.publish(EventType.UPLOADING, action.to_dict())
                self._set(AgentState.UPLOADING, "UPLOADING")
            else:
                self._set(
                    ACTION_STATE.get(action.type, AgentState.THINKING),
                    ACTION_DETAIL.get(action.type),
                )
            self._last_field = action.field_id or action.reason
            self._bus.publish(EventType.ACTION_STARTED, action.to_dict())
            result = self._executor.execute(action)
            results.append(result)
            if result.ok and is_upload:
                self._bus.publish(EventType.UPLOAD_COMPLETED, result.to_dict())
                self._snapshot("after-upload", index, record_key)
            if not result.ok and action.type in {ActionType.SUBMIT, ActionType.CLICK}:
                logger.warning("submit/click failed; stopping record: {}", result.message)
                break
        return results

    def _remaining_record(self, record: SourceRecord, handled_labels: set[str]) -> SourceRecord:
        """A copy of the source record minus the labels already written.

        The reveal pass maps only these onto freshly revealed fields, so each
        source value is typed exactly once into its own field instead of being
        re-bound to the next revealed control.
        """
        ordered = [label for label in record.ordered_labels if label not in handled_labels]
        return SourceRecord(
            pairs={label: record.pairs[label] for label in ordered if label in record.pairs},
            ordered_labels=ordered,
            title=record.title,
        )

    def _refresh_source_record(self, record: SourceRecord, scene: SceneDescription) -> SourceRecord:
        """Merge freshly-visible LEFT-panel source data into the working record.

        The left source panel scrolls together with the right entry form, so
        every scroll reveals more label:value rows. Without re-reading them,
        the fields revealed below the fold would have no value to fill. New
        labels are merged into the record; already-written labels are kept out.
        """
        try:
            pairs = self._collect_source_pairs(scene)
        except Exception as exc:
            logger.debug("left-panel source re-read failed: {}", exc)
            return record
        merged = dict(record.pairs)
        ordered = list(record.ordered_labels)
        changed = False
        for label, value in pairs:
            if label and label not in merged:
                merged[label] = value
                ordered.append(label)
                changed = True
        if not changed:
            return record
        return SourceRecord(pairs=merged, ordered_labels=ordered, title=record.title)

    def _scan_fill_revealed(
        self,
        record: SourceRecord,
        handled_ids: set[str],
        handled_labels: set[str] | None = None,
        submit_id: str | None = None,
        index: int = 0,
        initial_scene: SceneDescription | None = None,
    ) -> list[ActionResult]:
        """Fill-visible -> expand sections -> dual-panel scroll -> repeat.

        Treats the form as ONE continuous document with two independent
        scrollable panels (left source data + right entry form), exactly like a
        human operator:

        1. read    - re-read the LEFT source panel every round: each scroll
                     reveals more source rows whose values the newly revealed
                     fields below the fold need;
        2. scan    - fill every visible, unhandled field that has a value
                     (never skip a field that is on screen);
        3. expand  - click collapsible upload/attachment section headers so
                     their fields are revealed (never the submit button);
        4. scroll  - only once the current viewport is fully handled, scroll
                     BOTH panels a small incremental amount (click-focus inside
                     each panel first) and keep them synchronized;
        5. end     - stop ONLY when BOTH panels have reached their bottom
                     (i.e. no more content can be revealed on either side),
                     confirmed by the Upload Details section when present.

        The loop NEVER stops just because the current viewport is complete:
        it keeps scrolling as long as either panel still has content to reveal
        (live targets refresh their geometry after every scroll via
        ``field_map_refresh``). Bounded by ``max_scan_rounds``; never runs
        forever.
        """
        scroll_ctrl = DualPanelScroll(
            stall_limit=self._scroll_stall_limit,
            min_pixels=self._scroll_min_pixels,
            max_pixels=self._scroll_max_pixels,
            settle_range=self._scroll_settle,
        )
        rounds = max(1, self._max_scan_rounds)
        # Seed the scroll-progress baseline with the pre-reveal viewport: a
        # field that becomes visible on the FIRST reveal observation is already
        # scroll progress (the scroll before this pass moved it into view), so
        # the panel must count as "ever moved" from the very first round.
        if initial_scene is not None:
            scroll_ctrl.record_observation(initial_scene)
        # NEVER REVERSE SCROLL: the reveal pass is DOWN-only. While it runs,
        # the executor's scroll-into-view is forced DOWN so a failed scroll is
        # retried forward, never by scrolling back up.
        self._scroll_direction = "down"
        self._executor.set_scroll_direction("down")
        try:
            return self._scan_fill_revealed_rounds(
                record, handled_ids, handled_labels, submit_id, index, scroll_ctrl, rounds,
            )
        finally:
            self._executor.set_scroll_direction(None)
            self._scroll_direction = None

    def _scan_fill_revealed_rounds(
        self,
        record: SourceRecord,
        handled_ids: set[str],
        handled_labels: set[str] | None,
        submit_id: str | None,
        index: int,
        scroll_ctrl: DualPanelScroll,
        rounds: int,
    ) -> list[ActionResult]:
        results: list[ActionResult] = []
        for _round in range(rounds):
            if self._stop:
                break
            self._set(AgentState.OBSERVE_VIEWPORT)
            self._force_rebuild = True
            analysis, _ = self._observe()
            if analysis is None:
                break
            scene = analysis.scene
            viewport = self._viewport(analysis)

            # 1) read the LEFT source panel again: a scroll reveals more source
            #    rows, so lower fields always have a value to fill.
            self._set(AgentState.RECORD_EXTRACTION, "READING")
            record = self._refresh_source_record(record, scene)

            # 2) expand collapsed upload/section areas before filling.
            if self._expand_upload_section(scene, submit_id=submit_id, viewport=viewport):
                continue

            # 3) fill every visible, unhandled field that has a value.
            visible = self._visible_fields(discover_fields(scene), viewport)
            fresh = [f for f in visible if f.element_id not in handled_ids and f.bbox is not None]
            if fresh:
                # Only the source pairs not yet written to the form are mapped
                # onto the freshly revealed fields. Re-mapping the full record
                # would re-bind the first unfilled label (e.g. "Field 0") to
                # every new field revealed by a scroll and type its value again.
                handled_labels = handled_labels or set()
                pending = self._remaining_record(record, handled_labels)
                mapping = self._mapper.map(pending, fresh)
                sub_plan = self._planner.plan_fill(pending, mapping, scene, None)
                actionable = [a for a in sub_plan.actions if not self._is_submit_action(a, None)]
                if actionable:
                    self._lock_scrolling()
                    written_labels = {
                        m.target_id: m.source_label for m in mapping.mappings
                    }
                    for action in actionable:
                        if self._stop:
                            break
                        result = self._executor.execute(action)
                        results.append(result)
                        if action.field_id:
                            handled_ids.add(action.field_id)
                            if action.field_id in written_labels:
                                handled_labels.add(written_labels[action.field_id])
                    continue
                # These fields exist but have no value to fill: remember them so
                # we never re-scan them, then move on to scrolling.
                handled_ids.update(f.element_id for f in fresh)

            # 4) scroll permission: the LAST operation, only when the current
            #    viewport is complete (reveal-ready). Verification misses never
            #    block discovery (can_reveal_scroll, not can_scroll), so a
            #    single unverified value can never freeze the scan (Issue 1).
            if not self.can_reveal_scroll(scene, handled_ids, viewport, results, submit_id=submit_id):
                time.sleep(0.3)  # human-paced; bounded by max_scan_rounds
                continue

            # 5) track per-panel geometry + progress from the current scene so
            #    the completion checks know when a panel truly hit its bottom.
            #    When UIA scroll containers are available they are re-discovered
            #    and matched to the LEFT/RIGHT panels; their scroll percent then
            #    decides whether more content remains (never the viewport).
            self._refresh_scroll_session()
            containers = self._scroll_container_map(scroll_ctrl)
            scroll_ctrl.update_panels(self._panel_rects(scene), self._client_rect(), containers=containers)
            scroll_ctrl.record_observation(scene)

            # 6) completion: the spec's stop condition - Upload Details visible
            #    AND both panels reached their bottom AND no further scrolling
            #    is possible. Never stop just because the current viewport is
            #    complete - there are many fields below the fold on both sides.
            if scroll_ctrl.form_complete():
                logger.info("reveal pass done: {}", scroll_ctrl.completion_reason())
                break

            # 7) scroll BOTH panels a small incremental amount. When UIA
            #    containers are known each panel is scrolled *directly* through
            #    its own ScrollPattern (verified + retried), otherwise the prior
            #    click-focus wheel path is used. Panels stay synchronized.
            self._scroll_panels(analysis, scroll_ctrl)

        # 8) end of form: click submit/upload exactly once, never twice, and
        #    ONLY when the scan genuinely reached the bottom of both panels.
        #    A reveal pass that could not move any content (the MPF panels
        #    expose no ScrollPattern, so a failing scroll stalls without ever
        #    reaching the Upload Details section) must NOT submit - it would
        #    save a half-filled record. The failure is surfaced instead.
        #    Scrolling is unlocked BEFORE the click so the executor may scroll
        #    the button into view if it sits just below the fold, and it is
        #    ALWAYS released (finally) so the next record starts with full
        #    scroll freedom (the lock is re-engaged per record anyway).
        try:
            self._unlock_scrolling()
            if submit_id is not None:
                if scroll_ctrl.form_complete():
                    submit_result = self._click_submit_at_end(record, index=index)
                    if submit_result is not None:
                        results.append(submit_result)
                else:
                    results.append(ActionResult(
                        action=Action(
                            type=ActionType.STOP,
                            reason="form incomplete - Upload Details not reached",
                        ),
                        success=False,
                        verified=False,
                        message="form incomplete - Upload Details not reached; submit skipped",
                    ))
        finally:
            self._unlock_scrolling()
        return results

    def _click_submit_at_end(self, record: SourceRecord, index: int = 0) -> ActionResult | None:
        """Click the submit/upload button once after the form is fully handled.

        Re-observes so the click targets the current screen position (the loop
        may have scrolled since the button was first seen), then executes a
        single CLICK and surfaces the upload lifecycle events.
        """
        self._force_rebuild = True
        analysis, _ = self._observe()
        if analysis is None:
            return None
        scene = analysis.scene
        submit_id = self._find_submit(scene)
        if submit_id is None:
            return None
        element = scene.element(submit_id)
        if element is None or element.bbox is None:
            return None
        action = Action(
            type=ActionType.CLICK,
            field_id=submit_id,
            bbox=element.bbox.shifted(*scene.screen_offset),
            confidence=1.0,
            expected="clicked submit",
            reason="click submit button after filling the form",
        )
        self._bus.publish(EventType.UPLOADING, action.to_dict())
        self._set(AgentState.UPLOADING, "UPLOADING")
        self._last_field = action.field_id or action.reason
        result = self._executor.execute(action)
        if result.ok:
            self._bus.publish(EventType.UPLOAD_COMPLETED, result.to_dict())
            self._snapshot("after-upload", index, record.record_key or "")
        return result

    # -- viewport-driven scanning --------------------------------------------

    def _wait_until_stable(self, max_wait: float = 4.0, poll: float = 0.35) -> None:
        """Startup settle: wait for the attached UI to stop changing.

        Live targets re-render after attach (grids loading, panels painting);
        acting on a half-rendered form wastes the first viewport. Observations
        here are free and bounded, and the reveal pass guarantees no field is
        skipped. Stateless/mock targets skip the settle entirely.

        This phase is scroll-locked: nothing scrolls while the page settles.
        """
        if not self._settle_on_start or self._stop:
            return
        self._set(AgentState.OBSERVE_VIEWPORT)
        deadline = time.time() + max_wait
        last_sig = ""
        stable = 0
        while time.time() < deadline and not self._stop:
            self._force_rebuild = True
            analysis, _ = self._observe()
            if analysis is None or analysis.capture is None:
                break  # not a live capture target; nothing to settle
            sig = self._element_signature(analysis.scene)
            if sig and sig == last_sig:
                stable += 1
                if stable >= 2:
                    break
            else:
                stable = 0
            last_sig = sig
            time.sleep(poll)
        # Reset so the record-wait phase re-observes the settled screen fresh.
        self._cached_analysis = None
        self._last_signature = ""
        self._force_rebuild = False

    @staticmethod
    def _viewport(analysis: SceneAnalysis) -> tuple[int, int] | None:
        """The client-area viewport ``(width, height)``, or None when unknown.

        Vision targets only see the current viewport; the bbox of a below-fold
        field is outside it and must not be acted on until a scroll reveals it.
        """
        capture = analysis.capture
        if capture is None:
            return None
        width = getattr(capture, "width", 0) or 0
        height = getattr(capture, "height", 0) or 0
        if width <= 0 or height <= 0:
            return None
        return width, height

    @classmethod
    def _visible_fields(
        cls, fields: list[Any], viewport: tuple[int, int] | None
    ) -> list[Any]:
        """Fields whose band intersects the current viewport.

        Fields without a bbox (web-DOM controls) are always kept - visibility
        cannot be judged and skipping them would break pure-web targets. A
        field scrolled above the top edge (``bottom <= 0``) or fully below the
        fold (``top >= height``) is NOT visible yet.
        """
        if viewport is None:
            return fields
        _, height = viewport
        return [
            f
            for f in fields
            if f.bbox is None
            or (f.bbox.top < height and f.bbox.bottom > 0)
        ]

    @staticmethod
    def _is_submit_action(action: Action, submit_id: str | None) -> bool:
        return action.type == ActionType.SUBMIT or (
            action.type == ActionType.CLICK
            and submit_id is not None
            and action.field_id == submit_id
        )

    @staticmethod
    def _element_signature(scene: SceneDescription) -> str:
        """Order-independent snapshot of the scene geometry for change detection."""
        parts = []
        for e in scene.elements:
            if e.bbox is not None:
                parts.append(
                    f"{e.element_id}:{e.type.value}:{e.bbox.left},{e.bbox.top},{e.bbox.width},{e.bbox.height}"
                )
            else:
                parts.append(f"{e.element_id}:{e.type.value}:nobox")
        return "|".join(sorted(parts))

    def _expand_upload_section(
        self,
        scene: SceneDescription,
        submit_id: str | None = None,
        viewport: tuple[int, int] | None = None,
    ) -> bool:
        """Click a collapsed upload/attachment section header, if any.

        Expands the strongest candidate that is not the final submit button,
        is inside the current viewport, and has fields below it (i.e. it really
        is a section header, not a terminal action button). Returns True when a
        header was clicked so the caller re-observes.
        """
        for element in self._upload_section_candidates(scene, viewport, submit_id=submit_id):
            bbox = element.bbox
            if bbox is None:
                continue
            action = Action(
                type=ActionType.CLICK,
                field_id=element.element_id,
                bbox=bbox.shifted(*scene.screen_offset),
                confidence=element.confidence or 0.5,
                reason=f"expand '{element.label or element.name}' section",
            )
            result = self._executor.execute(action)
            if result.ok:
                self._expanded_sections.add(element.element_id)
                logger.info("expanded upload section '{}'", element.label or element.name)
                return True
        return False

    def _upload_section_candidates(
        self,
        scene: SceneDescription,
        viewport: tuple[int, int] | None,
        submit_id: str | None = None,
    ) -> list[ScreenElement]:
        """Expandable upload/attachment headers in the current viewport.

        Excludes the final submit/upload button and action-strip buttons (they
        are clicked once at the end, never expanded), and anything below the
        fold (the current viewport is handled before any scroll happens).
        """
        candidates: list[ScreenElement] = []
        for element in find_upload_sections(scene, exclude_ids=self._expanded_sections):
            if element.section == "actions":
                continue
            if submit_id is not None and element.element_id == submit_id:
                continue
            bbox = element.bbox
            if bbox is None:
                continue
            if viewport is not None and bbox.top >= viewport[1]:
                continue
            has_fields_below = any(
                f.bbox is not None and f.bbox.top >= bbox.bottom
                for f in discover_fields(scene)
            )
            if has_fields_below:
                candidates.append(element)
        return candidates

    def can_scroll(
        self,
        scene: SceneDescription,
        handled_ids: set[str],
        viewport: tuple[int, int] | None,
        results: list[ActionResult],
        submit_id: str | None = None,
    ) -> bool:
        """Scroll permission: True only when the current viewport is complete.

        Scrolling is the LAST operation for a viewport. It delegates to the
        :class:`ViewportModel` NO SCROLL RULE - every visible field handled,
        every dropdown/date done, uploads/attachments checked, and every value
        verified - before any scroll is permitted. The first blocked gate is
        surfaced in ``self._scroll_blocked_reason`` for the debug dump.
        """
        model = self._build_viewport_model(scene, handled_ids, viewport, results)
        ok = model.can_scroll
        self._scroll_blocked_reason = model.scroll_blocked_reason()
        self._write_viewport_debug(model)
        return ok

    def can_reveal_scroll(
        self,
        scene: SceneDescription,
        handled_ids: set[str],
        viewport: tuple[int, int] | None,
        results: list[ActionResult],
        submit_id: str | None = None,
    ) -> bool:
        """Reveal-pass scroll permission: like ``can_scroll`` but never blocked
        by a prior verification failure.

        The reveal pass exists to DISCOVER the fields still below the fold. A
        value that failed to verify on one field must not stop it from scrolling
        to reach the rest of the form (Issue 1) - otherwise a single miss freezes
        the scan and every lower field is skipped. Only unhandled VISIBLE
        controls (fields, dropdowns, dates, uploads) may block the scroll; every
        gate is exactly the ViewportModel NO SCROLL RULE minus the verification
        gate. Verification still runs per field; it just never halts discovery.
        """
        model = self._build_viewport_model(scene, handled_ids, viewport, results)
        ok = model.viewport_complete
        self._scroll_blocked_reason = None if ok else model.reveal_blocked_reason()
        self._write_viewport_debug(model)
        return ok

    def _build_viewport_model(
        self,
        scene: SceneDescription,
        handled_ids: set[str],
        viewport: tuple[int, int] | None,
        results: list[ActionResult],
    ) -> ViewportModel:
        return ViewportModel(
            scene=scene,
            viewport=viewport,
            handled_ids=set(handled_ids),
            expanded_upload_ids=self._expanded_sections,
            results=list(results),
            scroll_position=self._scroll_position,
        )

    def _lock_scrolling(self) -> None:
        """Freeze scrolling while the current viewport is observed/filled."""
        self._executor.set_scroll_allowed(lambda: False)

    def _unlock_scrolling(self) -> None:
        """Allow scroll-into-view again (reveal pass / final submit click)."""
        self._executor.set_scroll_allowed(lambda: True)

    def _panel_rects(self, scene: SceneDescription) -> dict[str, BBox | None]:
        """Absolute-screen regions of the two panels: left + right.

        Prefers the UIA field-map panel rects (left source list, right entry
        form), then any explicit region provider. When a provider yields two
        regions they are treated as [left, right]; a single region is treated
        as the entry form. The regions are later clipped to the visible client
        area so the cursor never lands below the fold.
        """
        rects: dict[str, BBox | None] = {PANEL_LEFT: None, PANEL_RIGHT: None}
        if self._field_map is not None:
            rects[PANEL_LEFT] = self._field_map.left_rect
            rects[PANEL_RIGHT] = self._field_map.right_rect
        if not any(rects.values()) and self._scroll_regions is not None:
            try:
                regions = [r for r in self._scroll_regions(scene) if r is not None]
            except Exception:
                regions = []
            if len(regions) >= 2:
                rects[PANEL_LEFT], rects[PANEL_RIGHT] = regions[0], regions[1]
            elif regions:
                rects[PANEL_RIGHT] = regions[0]
        return rects

    def _client_rect(self) -> tuple[int, int, int, int] | None:
        """Absolute client rect of the sandbox target, or None."""
        sandbox = getattr(self._executor, "_sandbox", None)
        if sandbox is None:
            return None
        try:
            target = sandbox.validate_target()
            if target is None or not target.client_rect:
                return None
            left, top, right, bottom = target.client_rect
            if right <= left or bottom <= top:
                return None
            return left, top, right, bottom
        except Exception:
            return None

    def _refresh_scroll_session(self) -> None:
        """Re-discover the UIA scroll containers for this observation round.

        The provider (wired by the desktop assistant) returns a fresh
        :class:`ScrollSession` whose containers carry the current scroll
        percent. No provider (tests / web / no UIA) means the reveal pass falls
        back to the click-focus wheel path.
        """
        if self._scroll_container_provider is None:
            self._scroll_session = None
            return
        try:
            session = self._scroll_container_provider()
        except Exception as exc:
            logger.debug("scroll container discovery failed: {}", exc)
            session = None
        self._scroll_session = session if (session is not None and session.available) else None

    def _scroll_container_map(self, scroll_ctrl: DualPanelScroll) -> dict[str, ScrollContainer | None]:
        """Match the discovered scroll containers onto the LEFT/RIGHT panels.

        Uses the field-map content rects (never hardcoded coordinates) so the
        source panel and entry panel each get their own scroll container.
        """
        result: dict[str, ScrollContainer | None] = {PANEL_LEFT: None, PANEL_RIGHT: None}
        if self._scroll_session is None:
            return result
        left_rect = self._field_map.left_rect if self._field_map is not None else None
        right_rect = self._field_map.right_rect if self._field_map is not None else None
        try:
            chosen = pick_left_right_containers(
                self._scroll_session.containers,
                left_rect,
                right_rect,
                self._client_rect(),
            )
        except Exception as exc:
            logger.debug("scroll container match failed: {}", exc)
            return result
        for name in (PANEL_LEFT, PANEL_RIGHT):
            result[name] = chosen.get(name)
        return result

    def _scroll_panels(self, analysis: SceneAnalysis, scroll_ctrl: DualPanelScroll) -> None:
        """Scroll BOTH panels one small incremental step, then settle.

        Called ONLY after ``can_reveal_scroll()`` has returned True (the
        current viewport is fully handled). Each panel is scrolled *directly*
        through its own discovered UIA scroll container (ScrollPattern, with
        verification + retry escalation) when available - the outer window and
        desktop are never scrolled. Without UIA containers the panel is located
        from the field map, click-focused INSIDE the panel and wheel-scrolled by
        the 250-350 px adaptive amount. If one panel moved while the other did
        not, the lagging panel is nudged again so the left source data and the
        right entry form always correspond to the same section. A human-paced
        settle (300-500 ms) lets the UI refresh before the next observation.
        """
        scene = analysis.scene
        index = getattr(analysis, "index", None)
        key = getattr(analysis, "record_key", None)
        self._snapshot("before-scroll", index or 0, key or "record")
        amount = scroll_ctrl.scroll_notches(scene)
        pixels = (scroll_ctrl.min_pixels + scroll_ctrl.max_pixels) // 2
        known = scroll_ctrl.known_panels()

        session = self._scroll_session
        if session is not None and session.scroller is not None:
            self._scroll_containers(session, scroll_ctrl, pixels, scene)
        elif not known:
            # No panel geometry: one human wheel scroll over the form.
            self._set(AgentState.SCROLLING, "SCROLLING FORM")
            self._executor.execute(Action(
                type=ActionType.SCROLL,
                value=None,
                scroll_amount=amount,
                reason="reveal next section of the form (dual-panel scroll unavailable)",
            ))
            self._scroll_position += amount * 50  # 1 wheel notch ~= 50px
        else:
            for name, panel in ((PANEL_LEFT, scroll_ctrl.left), (PANEL_RIGHT, scroll_ctrl.right)):
                if panel.rect is None or not scroll_ctrl.needs_scroll(name):
                    continue
                anchor = scroll_ctrl.scroll_anchor(name, scene)
                if anchor is None:
                    anchor = (panel.rect.center[0], panel.rect.top + 6)
                self._scroll_region(name, panel.rect, amount, anchor=anchor, reason=f"scroll {name} panel")
                panel.scroll_position += amount * 50
                self._scroll_position += amount * 50

            # Panel synchronization: if one side moved while the other did not
            # (and neither is at its bottom), nudge the lagging panel once more
            # before continuing so both sides stay in the same section.
            lagging = scroll_ctrl.lagging_panel()
            if lagging is not None:
                panel = scroll_ctrl.panel(lagging)
                if panel.rect is not None:
                    anchor = scroll_ctrl.scroll_anchor(lagging, scene)
                    if anchor is None:
                        anchor = (panel.rect.center[0], panel.rect.top + 6)
                    self._scroll_region(
                        lagging, panel.rect, amount, anchor=anchor, reason=f"resync {lagging} panel"
                    )
                    panel.scroll_position += amount * 50
                    self._scroll_position += amount * 50

        # Human-paced settle so the UI refreshes before the next observation.
        time.sleep(random.uniform(*scroll_ctrl.settle_range))
        self._snapshot("after-scroll", index or 0, key or "record")
        self._write_viewport_position(scene, scroll_ctrl)  # debug: current scroll position

    def _scroll_containers(
        self,
        session: ScrollSession,
        scroll_ctrl: DualPanelScroll,
        pixels: int,
        scene: SceneDescription,
    ) -> None:
        """Scroll every panel that still needs content via its UIA container.

        Each scroll is VERIFIED against that panel's own visible labels: a
        fresh observation is taken and the panel-scoped signature (element
        ids + types + visible labels + rects) must differ from the pre-scroll
        snapshot, otherwise the scroll failed and the :class:`PanelScroller`
        retries with the next method / a bigger distance. The lagging panel is
        re-scrolled so the two sides stay in the same section. A scroll that no
        longer moves anything updates ``more_content`` so the next round's
        bottom detection can finish.
        """
        before_signatures = {
            name: scroll_ctrl.panel_signature(scene, panel, include_labels=True)
            for name, panel in ((PANEL_LEFT, scroll_ctrl.left), (PANEL_RIGHT, scroll_ctrl.right))
            if panel.container is not None
        }

        def _verify(name: str) -> bool:
            self._set(AgentState.VERIFYING, "SCROLL VERIFY")
            refreshed = self.reobserve_scene()
            if refreshed is None:
                return False
            panel = scroll_ctrl.panel(name)
            before = before_signatures.get(name, "")
            return scroll_ctrl.panel_signature(refreshed, panel, include_labels=True) != before

        for name, panel in ((PANEL_LEFT, scroll_ctrl.left), (PANEL_RIGHT, scroll_ctrl.right)):
            container = panel.container
            if container is None or not scroll_ctrl.needs_scroll(name):
                continue
            self._scroll_one_container(session, panel, container, pixels, lambda name=name: _verify(name), name)
        lagging = scroll_ctrl.lagging_panel()
        if lagging is not None:
            panel = scroll_ctrl.panel(lagging)
            container = panel.container
            if container is not None:
                self._scroll_one_container(session, panel, container, pixels, lambda: _verify(lagging), lagging)

    def _scroll_one_container(
        self,
        session: ScrollSession,
        panel: Any,
        container: ScrollContainer,
        pixels: int,
        verify: Callable[[], bool],
        name: str,
    ) -> None:
        if session.scroller is None:
            return
        self._set(AgentState.SCROLLING, f"SCROLLING {name.upper()} PANEL")
        try:
            outcome = session.scroller.scroll_down(container, pixels, verify)
        except Exception as exc:
            logger.debug("container scroll failed for {} panel: {}", name, exc)
            outcome = None
        if outcome is not None:
            panel.scroll_position += pixels
            self._scroll_position += pixels
            panel.more_content = container.more_content
            method = "FAILED" if outcome.method == "none" else outcome.method.upper()
            self._set(AgentState.SCROLLING, f"SCROLLING {name.upper()} PANEL ({method})")
            logger.info(
                "scrolled {} panel via '{}' (changed={}, percent={})",
                name,
                outcome.method,
                outcome.changed,
                container.vertical_scroll_percent,
            )
        succeeded = outcome is not None and outcome.method != "none"
        if succeeded:
            self._panel_scroll_failures[name] = 0
            return
        failures = self._panel_scroll_failures.get(name, 0) + 1
        self._panel_scroll_failures[name] = failures
        if failures < _RAW_SCROLL_FAILSAFE_THRESHOLD or container.rect is None:
            return
        # Every structured method (UIA ScrollPattern, mouse wheel, scrollbar
        # drag, keyboard, plugin override) has now failed to move this panel
        # for `_RAW_SCROLL_FAILSAFE_THRESHOLD` cycles in a row. Rather than
        # let the panel go permanently idle (the exact "it will not scroll
        # and just sits there" failure mode), force the simplest possible
        # thing that can still work: click into the panel and send a big,
        # unconditional wheel scroll directly, bypassing PanelScroller
        # entirely. The counter is reset either way so this only fires again
        # after another full run of consecutive failures, not every cycle.
        logger.error(
            "{} panel scroll stuck after {} consecutive failures - forcing raw wheel fallback",
            name,
            failures,
        )
        self._set(AgentState.RECOVERY, f"SCROLLING {name.upper()} PANEL (FORCED FALLBACK)")
        anchor = (container.rect.center[0], container.rect.top + 6)
        self._scroll_region(
            name,
            container.rect,
            max(pixels * 3, self._scroll_max_pixels * 2),
            anchor=anchor,
            reason=f"forced raw wheel fallback for stuck {name} panel",
        )
        self._panel_scroll_failures[name] = 0

    def _scroll_region(
        self,
        name: str,
        region: BBox,
        amount: int,
        anchor: tuple[int, int] | None = None,
        reason: str = "scroll region",
    ) -> None:
        """Move the cursor inside the panel, CLICK to focus it, then wheel-scroll.

        A wheel event scrolls whichever pane sits under the cursor, so each
        panel of a split form is scrolled separately. The panel is click-focused
        FIRST (a human clicks inside the panel before scrolling) so the wheel
        is captured by THIS panel's own scrollbar - never the whole window.
        ``anchor`` is a safe point inside the panel (over a label / header), and
        the cursor is clamped to the visible client area so it never lands below
        the fold.
        """
        if anchor is None:
            cx = max(region.left + 4, min(region.right - 4, int(region.center[0])))
            cy = max(region.top + 4, min(region.bottom - 4, int(region.center[1])))
        else:
            cx = max(region.left + 4, min(region.right - 4, int(anchor[0])))
            cy = max(region.top + 4, min(region.bottom - 4, int(anchor[1])))
        self._set(AgentState.SCROLLING, f"SCROLLING {name.upper()} PANEL (MOUSE WHEEL)")
        try:
            self._mouse_move_to(cx, cy)
        except Exception as exc:
            logger.debug("move to scroll region failed: {}", exc)
        # Click INSIDE the panel to focus it so the wheel scrolls this panel.
        self._executor.execute(Action(
            type=ActionType.CLICK,
            value=None,
            field_id=None,
            bbox=BBox(cx - 1, cy - 1, 2, 2),
            reason=f"focus panel for scroll at ({cx},{cy})",
        ))
        self._executor.execute(Action(
            type=ActionType.SCROLL,
            value=None,
            scroll_amount=amount,
            reason=reason,
        ))

    def _mouse_move_to(self, x: int, y: int) -> None:
        """Best-effort human mouse move to absolute screen coords (if available)."""
        mouse = getattr(self._executor, "_mouse", None)
        if mouse is not None:
            mouse.move_to(x, y)

    # -- debug dumps ----------------------------------------------------------

    def _debug_path(self, name: str) -> Path | None:
        if self._debug_dir is None:
            return None
        path = self._debug_dir / name
        try:
            path.parent.mkdir(parents=True, exist_ok=True)
        except Exception:
            return None
        return path

    def _write_debug(self, name: str, data: Any) -> None:
        path = self._debug_path(name)
        if path is None:
            return
        try:
            path.write_text(json.dumps(data, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
        except Exception as exc:
            logger.debug("failed to write {}: {}", path, exc)

    def _dump_run_metrics(self, summary: WorkflowSummary) -> None:
        """Write per-field stage timings to ``debug/performance/run_metrics.json``.

        The spec's per-stage budget split (discovery/action/verify/ocr/recovery)
        is auditable here: each verifyable field's action/verify/recovery times
        accumulate, so a field eating 70-80s in retries is immediately visible.
        """
        if self._debug_dir is None:
            return
        try:
            metrics = self._executor.field_metrics() if self._executor is not None else {}
        except Exception as exc:
            logger.debug("field metrics unavailable: {}", exc)
            metrics = {}
        if not metrics:
            return
        folder = self._debug_dir / "performance"
        try:
            folder.mkdir(parents=True, exist_ok=True)
        except Exception:
            return
        totals: dict[str, float] = {}
        for entry in metrics.values():
            for stage, seconds in entry.get("stages", {}).items():
                totals[stage] = totals.get(stage, 0.0) + float(seconds)
        try:
            (folder / "run_metrics.json").write_text(
                json.dumps({
                    "run": {
                        "completed": summary.completed,
                        "failed": summary.failed,
                        "unverified_records": summary.unverified,
                        "unverified_fields": summary.unverified_fields,
                        "stopped_reason": summary.stopped_reason,
                    },
                    "fields": metrics,
                    "totals_by_stage": totals,
                }, ensure_ascii=False, indent=2, default=str),
                encoding="utf-8",
            )
        except Exception as exc:
            logger.debug("failed to write run_metrics.json: {}", exc)

    def _dump_verification_debug(self, summary: WorkflowSummary) -> None:
        """Write every structured verification event for post-mortem analysis.

        Each event carries expected/observed, the strategy evidence and the
        window geometry at read time - so a "vision read empty" can be traced
        to a genuinely empty field vs. a stale bbox (window resized/scrolled
        between the write and the read).
        """
        if self._debug_dir is None:
            return
        events = self._executor.verification_events() if self._executor is not None else []
        if not events:
            return
        self._write_debug("verification_debug.json", {
            "run": {
                "completed": summary.completed,
                "failed": summary.failed,
                "unverified_records": summary.unverified,
                "unverified_fields": summary.unverified_fields,
                "stopped_reason": summary.stopped_reason,
            },
            "verifications": events,
            "summary": {
                "total": len(events),
                "matched": sum(1 for e in events if e.get("ok")),
                "mismatched": sum(1 for e in events if not e.get("ok")),
                "unknown": sum(1 for e in events if e.get("status") == "UNKNOWN"),
            },
        })

    def _write_viewport_debug(self, model: ViewportModel) -> None:
        """Emit the current viewport model the last time scrolling was weighed.

        Let an operator audit exactly what the agent saw and why it did or did
        not scroll (every NO SCROLL RULE gate, every pending field).
        """
        if self._debug_dir is None:
            return
        try:
            payload = model.to_dict()
            payload["scroll_blocked_reason"] = self._scroll_blocked_reason
            self._write_debug("viewport.json", payload)
        except Exception as exc:
            logger.debug("viewport debug write failed: {}", exc)

    def _write_viewport_position(self, scene: SceneDescription, scroll_ctrl: DualPanelScroll) -> None:
        """Persist the current scroll offset for both panels while scanning (debug)."""
        if self._debug_dir is None:
            return
        data = {
            "scroll_position": self._scroll_position,
            "window_title": scene.window_title,
            "layout_summary": scene.layout_summary,
            "panels": scroll_ctrl.to_dict().get("panels", {}),
            "upload_visible": scroll_ctrl.upload_visible,
        }
        self._write_debug("scroll_position.json", data)

    def _dump_record_debug(
        self, plan: FillPlan, results: list[ActionResult], index: int, record: SourceRecord
    ) -> None:
        if self._debug_dir is None:
            return
        self._write_session("record.json", {
            "index": index,
            "key": record.record_key,
            "title": record.title,
            "pairs": dict(record.pairs),
            "ordered_labels": record.ordered_labels,
        })
        self._write_debug("planner.json", plan.to_dict())
        self._write_debug("execution_plan.json", plan.to_dict())
        self._write_debug("execution.json", {
            "record_index": index,
            "actions": [r.to_dict() for r in results],
        })
        self._write_debug("verification.json", {
            "record_index": index,
            "results": [
                {
                    "field_id": r.action.field_id,
                    "action": r.action.type.value,
                    "expected": r.action.expected or r.action.value,
                    "verified": r.verified,
                    "success": r.success,
                    "evidence": r.verification_evidence or r.message,
                }
                for r in results
            ],
        })

    def _write_session(self, name: str, payload: dict) -> Path | None:
        if self._session_dir is None:
            return None
        try:
            self._session_dir.mkdir(parents=True, exist_ok=True)
            path = self._session_dir / name
            path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
            return path
        except Exception as exc:
            logger.debug("session write failed: {}", exc)
            return None

    def _snapshot(self, context: str, index: int, key: str) -> Path | None:
        """Capture a screenshot for the given record lifecycle point.

        Writes ``debug/screenshots/{index}-{key}-{context}.png`` via the
        target-agnostic capture callback. Never raises.
        """
        if self._capture_callback is None or self._debug_dir is None:
            return None
        folder = self._debug_dir / "screenshots"
        try:
            folder.mkdir(parents=True, exist_ok=True)
        except Exception:
            return None
        safe_key = "".join(c if c.isalnum() or c in "-_" else "_" for c in (key or "record"))
        path = folder / f"{index:04d}-{safe_key}-{context}.png"
        try:
            if self._capture_callback(path):
                log_screenshot(path, context)
                return path
        except Exception as exc:
            logger.debug("screenshot {} failed: {}", context, exc)
        return None

    def _dump_timeline(self, summary: WorkflowSummary) -> None:
        if self._session_dir is None:
            return
        self._write_session("timeline.json", {
            "records": summary.completed,
            "failed": summary.failed,
            "unverified_records": summary.unverified,
            "unverified_fields": summary.unverified_fields,
            "stopped_reason": summary.stopped_reason,
            "duration_ms": summary.total_duration * 1000.0,
            "records_json": [r.to_dict() for r in summary.records],
        })

    def _dump_failure(self, summary: WorkflowSummary) -> None:
        """Write ``failure.json`` when the run did not fully succeed.

        A clean ``max_records`` stop is not a failure; an aborted run (no
        record, stopped early with pending records) is.
        """
        if self._debug_dir is None:
            return
        failed = summary.failed
        clean_stop = not summary.stopped_reason or summary.stopped_reason.startswith("max_records")
        if failed == 0 and clean_stop:
            return
        payload = {
            "failed_records": failed,
            "completed_records": summary.completed,
            "unverified_records": summary.unverified,
            "unverified_fields": summary.unverified_fields,
            "stopped_reason": summary.stopped_reason,
            "duration_ms": summary.total_duration * 1000.0,
            "records": [r.to_dict() for r in summary.records],
            "state": self._states.state.value,
            "current_field": self._last_field,
            "planner_status": self._planner_status,
            "last_exception": self._last_exception,
            "no_record_reason": self._no_record_last_reason,
        }
        self._write_debug("failure.json", payload)

    def _dump_focus_history(self) -> None:
        """Write ``focus_history.json`` from the RECOVERY event stream.

        Every focus-related pause/refocus decision published by the sandbox or
        the workflow is replayed here so focus-loss episodes are auditable
        offline without re-running the automation.
        """
        if self._debug_dir is None:
            return
        try:
            history = [
                e.to_dict()
                for e in self._bus.history(EventType.RECOVERY)
                if "focus" in str(e.data.get("reason", "")).lower()
            ]
            self._write_debug("focus_history.json", {
                "count": len(history),
                "events": history,
            })
        except Exception as exc:
            logger.debug("focus_history write failed: {}", exc)

    def _dump_watchdog(self) -> None:
        """Write ``watchdog.json`` summarising both watchdog levels.

        Level 1 (sandbox focus/recovery events) and level 2 (state-budget
        overruns) are consolidated so a stuck run is diagnosable offline: which
        states overran, how often each was repeated, and what focus recovery
        happened around them.
        """
        if self._debug_dir is None:
            return
        try:
            recovery_events = [
                e.to_dict()
                for e in self._bus.history(EventType.RECOVERY)
            ]
            overruns = {
                str(state.value): count
                for state, count in self._state_overruns.items()
                if count > 0
            }
            self._write_debug("watchdog.json", {
                "level1_focus_events": [
                    e for e in recovery_events
                    if "focus" in str(e.get("data", {}).get("reason", "")).lower()
                ],
                "level2_state_overruns": overruns,
                "state_overrun_total": sum(overruns.values()),
                "recovery_events_total": len(recovery_events),
            })
        except Exception as exc:
            logger.debug("watchdog write failed: {}", exc)

    # -- helpers --------------------------------------------------------------

    def _await_record(self, previous_key: str | None) -> tuple[SceneAnalysis, SourceRecord] | None:
        """Wait for the next source record.

        Event-driven: the screen model is only rebuilt when the observed scene
        actually changes (app switch, upload click, left-panel change, scroll or
        focused-control change). On ``records == 0`` the loop never terminates:
        it shows "No valid record detected.", keeps retrying and waits for the
        next record. It only returns ``None`` when the loop is stopped.
        """
        self._set(AgentState.WATCHING)
        while not self._stop:
            self._check_state_budget()
            if self._next_timeout is not None and not self._stop:
                deadline = time.time() + self._next_timeout
                while not self._stop and time.time() < deadline:
                    if self._target is not None and not self._target.is_alive():
                        logger.warning("target window disappeared while awaiting a record")
                        self._set(AgentState.STOPPED)
                        return None
                    # UIA-first (cheap, no VLM): detect the next record straight
                    # from the left source panel before spending a full observe.
                    uia_record = self._read_source_uia_only()
                    if uia_record is not None:
                        if self._accept_record(uia_record, previous_key, None):
                            self._set(AgentState.REOBSERVE, "REOBSERVE")
                            self._force_rebuild = True
                            analysis, _ = self._observe()
                            if analysis is None:
                                continue
                            record = self._extract_record(analysis.scene)
                            if record is not None and self._accept_record(record, previous_key, analysis.scene):
                                self._set(AgentState.NEXT_RECORD, "NEXT_RECORD")
                                self._log_reset_transition(previous_key, record)
                                self._bus.publish(EventType.NEXT_RECORD_DETECTED, {"key": record.record_key})
                                return analysis, record
                        time.sleep(self._next_poll)
                        continue
                    analysis, changed = self._observe()
                    if analysis is not None and changed:
                        record = self._extract_record(analysis.scene)
                        if record is not None and self._accept_record(record, previous_key, analysis.scene):
                            self._log_reset_transition(previous_key, record)
                            self._bus.publish(EventType.NEXT_RECORD_DETECTED, {"key": record.record_key})
                            return analysis, record
                        if record is None:
                            # A no-record (e.g. loading) screen must not be
                            # cached by signature: force a fresh observation on
                            # the next poll so the following record is detected.
                            self._force_rebuild = True
                            time.sleep(self._next_poll)
                            continue
                        if self._same_record(record, previous_key):
                            # Same record: never trust the cached screen model -
                            # force a rebuild on the next poll so a submit/reset
                            # happening behind the scenes is still detected.
                            self._force_rebuild = True
                            self._bus.publish(EventType.NEXT_RECORD_WAITING, {"key": record.record_key})
                    time.sleep(self._next_poll)
                if not self._stop:
                    self._report_no_record()
                    if self._no_record_reason_code in HARD_FAILURE_CODES:
                        # A genuine, visible source panel that can never be
                        # read (no region / capture / OCR text / VLM, or every
                        # pair empty). Stop cleanly with the exact reason
                        # instead of spinning for minutes on a record that
                        # will never arrive (the "no record: no valid record
                        # detected" loop with BATCH COMPLETE: 0 record(s)).
                        logger.warning(
                            "no source record readable - stopping cleanly [{}]",
                            self._no_record_reason_code,
                        )
                        self._terminate_requested = True
                        self._terminate_reason_code = self._no_record_reason_code
                        self._stop = True
                        return None
            else:
                analysis, changed = self._observe()
                if analysis is not None and changed:
                    record = self._extract_record(analysis.scene)
                    if record is not None and self._accept_record(record, previous_key, analysis.scene):
                        self._log_reset_transition(previous_key, record)
                        self._bus.publish(EventType.NEXT_RECORD_DETECTED, {"key": record.record_key})
                        return analysis, record
                    if record is None:
                        # Never let a no-record screen stay cached (see above).
                        self._force_rebuild = True
                time.sleep(self._next_poll)
        self._set(AgentState.STOPPED)
        return None

    def _accept_record(self, record: SourceRecord, previous_key: str | None, scene: SceneDescription) -> bool:
        key = record.record_key
        if key and key != previous_key:
            return True
        valued = [v for v in record.pairs.values() if str(v).strip()]
        if not valued:
            return False
        if previous_key is None and len(valued) >= self._source_min_valued_pairs:
            # FIRST record: a partial read with enough valued pairs counts even
            # without a readable record key (the visual source reader may not
            # see an App No row). This is the fix for "no record: no valid
            # record detected" turning into BATCH COMPLETE: 0 record(s).
            return True
        if key is None and scene is not None and scene.layout_summary != (self._last_layout or ""):
            self._last_layout = scene.layout_summary
            return True
        return False

    def _same_record(self, record: SourceRecord, previous_key: str | None) -> bool:
        key = record.record_key
        if key and key == previous_key:
            return True
        return bool(key is None and self._last_layout and record.record_key is None)

    def reobserve_scene(self) -> SceneDescription | None:
        """Force a fresh observation and return the (field-map-merged) scene.

        Used by the executor after scrolling so bboxes stay accurate. Never
        raises: returns None on observation failure.
        """
        try:
            self._force_rebuild = True
            analysis, _ = self._observe()
            return analysis.scene if analysis is not None else None
        except Exception as exc:
            logger.debug("reobserve_scene failed: {}", exc)
            return None

    def _observe(self) -> tuple[SceneAnalysis | None, bool]:
        """Observe the target and rebuild the screen model only when changed.

        Returns ``(analysis, changed)``. ``changed`` is True when the screen
        model was rebuilt (first observation, app change, scroll, focus change,
        upload click or forced rebuild).
        """
        self._set(AgentState.OBSERVING)
        signature = self._target.signature() if hasattr(self._target, "signature") else ""
        if self._cached_analysis is None or self._force_rebuild or signature != self._last_signature:
            self._last_signature = signature
            self._force_rebuild = False
            # Live desktop forms refresh their UIA geometry after a scroll (the
            # below-fold fields move into view). Re-merge the current rects so
            # the reveal pass keeps discovering newly visible fields instead of
            # working against stale attach-time geometry.
            self._refresh_field_map_once()
            try:
                analysis = self._target.observe()
            except Exception as exc:
                # Vision/network failure (timeout, connection error, malformed
                # provider response) must never crash the whole run - fall
                # back to the last good observation (if any) and let the
                # normal polling/retry loop try again next cycle. This is the
                # concrete fix for "ATLAS must not crash merely because the
                # AI provider is unavailable."
                logger.warning("observation failed ({}); reusing last known screen state", exc)
                if self._cached_analysis is not None:
                    return self._cached_analysis, False
                return None, False
            if analysis is None:
                return None, False
            if self._scene_hook is not None:
                try:
                    analysis.scene = self._scene_hook(analysis.scene)
                except Exception:
                    logger.exception("scene hook failed; using raw scene")
            self._merge_field_map(analysis.scene)
            self._bus.publish(EventType.OBSERVED, analysis.to_dict())
            self._cached_analysis = analysis
            self._write_debug("vision_output.json", {
                "provider": analysis.scene.provider,
                "window_title": analysis.scene.window_title,
                "layout_summary": analysis.scene.layout_summary,
                "screen_offset": list(analysis.scene.screen_offset),
                "sections": [s.to_dict() for s in analysis.scene.sections],
                "elements": [e.to_dict() for e in analysis.scene.elements],
            })
            return analysis, True
        return self._cached_analysis, False

    def _extract_record(self, scene: SceneDescription) -> SourceRecord | None:
        """Run the Record Extraction stage from UIA/OCR source pairs, falling
        back to the VLM scene reader. On failure writes ``debug/no_record.json``
        and ``debug/record_failure.json`` (Step 6).
        """
        self._set(AgentState.RECORD_EXTRACTION)
        pairs = self._collect_source_pairs(scene)
        result = self._record_builder.build(pairs, title=scene.window_title)
        if result.record is None:
            self._report_no_record(scene, result)
            self._write_record_failure(scene, result)
            return None
        self._bus.publish(EventType.SOURCE_READ, result.record.to_dict())
        return result.record

    def _write_record_failure(self, scene: SceneDescription, result: RecordBuildResult) -> None:
        """Write ``debug/record_failure.json`` with full diagnostics (Step 6)."""
        if self._debug_dir is None:
            return
        payload = {
            "reason": result.reason or "record could not be built",
            "detected_labels": list(result.labels),
            "detected_values": list(result.values),
            "missing_required": list(result.missing_required),
            "missing_controls": [],
            "missing_mappings": [],
            "window_title": scene.window_title,
            "layout_summary": scene.layout_summary,
        }
        # Report missing controls/mappings from the field map if available.
        if self._field_map is not None:
            payload["missing_controls"] = [
                n.name for n in (self._field_map.right_fields or [])
                if n.control_type in {"Edit", "ComboBox", "CheckBox", "RadioButton"}
            ]
            payload["missing_mappings"] = [
                m for m in (self._field_map.mappings or [])
            ]
        self._write_debug("record_failure.json", payload)

    @staticmethod
    def _clean_node_name(node: Any) -> str:
        return (node.name or node.automation_id or "").strip()

    @staticmethod
    def _merge_source_pairs(
        *sources: list[tuple[str, str]],
    ) -> list[tuple[str, str]]:
        """Merge multiple (label, value) sources with priority
        ``sources[0] > sources[1] > ...`` (i.e. UIA before OCR before Vision).

        Per label, the first source with a NON-EMPTY value wins; an empty
        value from a higher-priority source never overrides a non-empty value
        found by a lower-priority one - this is the concrete fix for "OCR
        partial results must not destroy UIA results" and vice versa (PHASE 3).
        Label order follows first-seen order across all sources.
        """
        merged: dict[str, str] = {}
        ordered: list[str] = []
        for source_pairs in sources:
            for label, value in source_pairs:
                value = (value or "").strip()
                if label not in merged:
                    merged[label] = value
                    ordered.append(label)
                elif not merged[label] and value:
                    merged[label] = value
        return [(label, merged[label]) for label in ordered]

    def _write_source_pairs_debug(
        self,
        uia_pairs: list[tuple[str, str]],
        ocr_pairs: list[tuple[str, str]],
        merged: list[tuple[str, str]],
        diagnostics: PairingDiagnostics | None,
        left_label_count: int,
    ) -> None:
        """``debug/mpf/source_pairs.json`` (PHASE 2 diagnostic requirement)."""
        if self._debug_dir is None:
            return
        non_empty = sum(1 for _, v in merged if v)
        payload: dict[str, Any] = {
            "left_labels": left_label_count,
            "uia_pairs": len(uia_pairs),
            "ocr_pairs": len(ocr_pairs),
            "merged_pairs": len(merged),
            "non_empty_values": non_empty,
            "empty_values": len(merged) - non_empty,
            "pairs": [{"label": label, "value": value} for label, value in merged],
        }
        if diagnostics is not None:
            payload["pairing_diagnostics"] = diagnostics.to_dict()
        try:
            out_dir = self._debug_dir / "mpf"
            out_dir.mkdir(parents=True, exist_ok=True)
            (out_dir / "source_pairs.json").write_text(
                json.dumps(payload, ensure_ascii=False, indent=2, default=str), encoding="utf-8"
            )
        except Exception as exc:
            logger.debug("failed to write debug/mpf/source_pairs.json: {}", exc)
        logger.info(
            "[SOURCE] left labels={} candidate rows={} paired rows={} non-empty values={} empty values={}",
            left_label_count,
            diagnostics.to_dict()["candidate_rows"] if diagnostics is not None else "n/a",
            len(merged),
            non_empty,
            len(merged) - non_empty,
        )

    def _collect_source_pairs(self, scene: SceneDescription) -> list[tuple[str, str]]:
        """Merge UIA + OCR (+ existing VLM scene) source pairs with priority
        UIA -> OCR -> Vision, preferring whichever has a non-empty value per
        label (PHASE 3/4 fix) - a partial OCR read can no longer overwrite a
        good UIA value, and a UIA row that failed to pair (empty value) is now
        rescued by OCR instead of the whole record being rejected outright.

        When the merged UIA+OCR result is still entirely empty, the visual
        source observer (crop -> OCR -> VLM on the LEFT panel) is consulted as
        the image-based last resort, so a record is read even when UIA exposes
        no label/value rows at all.
        """
        uia_pairs: list[tuple[str, str]] = []
        labels: list[Any] = []
        self._last_source_reason_code: str = ""
        has_map = self._field_map is not None and self._field_map.has_source
        if has_map and self._field_map is not None and self._field_map.left_labels:
            labels = self._field_map.left_labels
            diagnostics = PairingDiagnostics()
            # UIA-only geometric pairing first (cheapest, most reliable
            # when it succeeds - live MPF exposes each value as a sibling
            # text node).
            uia_pairs = pair_source_pairs([], labels, diagnostics=diagnostics, member_only=True)

            ocr_pairs: list[tuple[str, str]] = []
            if self._ocr_callback is not None and self._field_map.left_rect is not None:
                left = self._field_map.left_rect
                try:
                    lines = self._ocr_callback(BBox(left.left, left.top, left.width, left.height))
                except Exception as exc:
                    logger.debug("source OCR failed: {}", exc)
                    lines = []
                self._write_debug("ocr_output.json", {
                    "region": {"left": left.left, "top": left.top, "width": left.width, "height": left.height},
                    "lines": [line.to_dict() for line in lines],
                })
                if lines:
                    ocr_pairs = pair_source_pairs(lines, labels, member_only=True)

            merged = self._merge_source_pairs(uia_pairs, ocr_pairs)
            if not merged:
                # Neither UIA row-pairing nor OCR found ANY row at all -
                # last resort: emit every known label with an empty value
                # so the record builder's diagnostics show exactly which
                # labels were seen (rather than silently returning []).
                merged = [
                    (label, "")
                    for label in map(self._clean_node_name, labels)
                    if is_member_field(label)
                ]
            self._write_source_pairs_debug(uia_pairs, ocr_pairs, merged, diagnostics, len(labels))
            if any(v for _, v in merged):
                return merged
            # All values empty even after UIA+OCR merge: fall through to the
            # image-based source observer rather than reporting "no record" on
            # the first attempt.

        # Visual source panel read (crop -> OCR -> VLM) - reads the record from
        # the LEFT panel IMAGE. This is the fix for "no record: no valid record
        # detected" that occurred whenever UIA pairing produced no values.
        if self._source_observer is not None:
            try:
                obs = self._source_observer.observe(
                    uia_pairs=uia_pairs,
                    known_labels=labels or None,
                    left_rect=(
                        self._field_map.left_rect
                        if self._field_map is not None and self._field_map.left_rect is not None
                        else None
                    ),
                )
            except Exception as exc:
                logger.debug("source observer failed: {}", exc)
                obs = None
            if obs is not None:
                self._write_debug("source_observer.json", obs.to_dict())
                if not obs.ok:
                    logger.debug("source observer reason: {}", obs.error_reason)
                    self._last_source_reason_code = obs.error_reason or ""
                elif obs.pairs:
                    merged = self._merge_source_pairs(obs.pairs, [])
                    if any(v for _, v in merged):
                        return merged
                    if len(merged) >= self._source_min_valued_pairs:
                        return merged
                    logger.debug(
                        "source observer returned {} valued pair(s) (< {}): not a valid record",
                        sum(1 for _, v in merged if v),
                        self._source_min_valued_pairs,
                    )
        record = self._source_reader.read(scene)
        return [(label, record.pairs.get(label, "")) for label in record.ordered_labels]

    def _read_source_uia_only(self) -> SourceRecord | None:
        """Read the left source panel using ONLY UIA (no VLM, no OCR, and -
        this is the fix - no full field-map rebuild on every poll).

        Live MPF exposes the source label/value rows as sibling UIA text
        nodes AND (being a classic WinForms app) each label is usually a real
        native control with its own HWND, so refreshing its text is a single
        cheap ``WM_GETTEXT``-style read (``UiaBackend.refresh_source_values``)
        rather than a full UIA tree rebuild. The previous version called
        ``_refresh_field_map_once()`` - the same expensive rebuild used for
        scroll/dropdown/verification events - on every single poll while
        waiting for a record, which is the root cause of "uia map built"
        repeating every ~9-10s for minutes with zero progress.

        The expensive rebuild is now used only as a throttled fallback (at
        most once every ``_source_refresh_interval`` seconds) for the case
        where none of the cached label nodes have a handle to cheap-read.
        Returns None while the left panel is loading/cleared, so callers keep
        waiting.
        """
        if self._field_map_refresh is None or self._field_map is None or not self._field_map.has_source:
            return None
        field_map = self._field_map
        if not field_map.left_labels:
            return None

        refreshed_labels = field_map.left_labels
        cheap_reads = 0
        try:
            from atlas.observe.uia import UiaBackend
            refreshed_labels, cheap_reads = UiaBackend.instance().refresh_source_values(field_map.left_labels)
        except Exception as exc:
            logger.debug("cheap source value read failed: {}", exc)

        if cheap_reads == 0:
            now = time.time()
            if now - self._last_source_full_refresh >= self._source_refresh_interval:
                self._last_source_full_refresh = now
                try:
                    self._refresh_field_map_once()
                except Exception as exc:
                    logger.debug("uia source refresh failed: {}", exc)
                field_map = self._field_map
                if field_map is None or not field_map.has_source or not field_map.left_labels:
                    return None
                refreshed_labels = field_map.left_labels
            # else: nothing cheap-readable and not due for a full refresh yet
            # - use whatever values were last known (still lets the loop keep
            # polling cheaply instead of blocking on a rebuild every cycle).

        try:
            pairs = pair_source_pairs([], refreshed_labels, member_only=True)
        except Exception as exc:
            logger.debug("uia-only source pairing failed: {}", exc)
            pairs = []
        if not pairs:
            return None
        try:
            result = self._record_builder.build(pairs, title="")
        except Exception as exc:
            logger.debug("uia-only record build failed: {}", exc)
            return None
        return result.record

    def _report_no_record(self, scene: SceneDescription | None = None, result: RecordBuildResult | None = None) -> None:
        """Surface a no-record condition and write ``debug/no_record.json``.

        Always assigns an exact reason code (NO_PAIRS_DETECTED,
        ALL_PAIRS_EMPTY, NO_VALID_RECORD, or a source-observer code such as
        SOURCE_NOT_FOUND / VISION_FAILED / NO_TEXT_DETECTED) so a 0-record run
        is root-caused from the diagnostics instead of a generic message.
        """
        self._set(AgentState.WAITING)
        reason = (result.reason if result is not None else None) or "no valid record detected"
        code = self._reason_code_for(reason)
        # A hard failure from the visual source observer (SOURCE_NOT_FOUND /
        # CAPTURE_FAILED / VISION_FAILED / NO_TEXT_DETECTED / ALL_PAIRS_EMPTY)
        # is more precise than the generic build-failure mapping, and it is the
        # one that triggers the clean self-termination in ``_await_record``.
        if getattr(self, "_last_source_reason_code", "") in HARD_FAILURE_CODES:
            code = self._last_source_reason_code
        self._no_record_reason_code = code
        if reason != self._no_record_last_reason:
            self._no_record_last_reason = reason
            logger.warning("no record: {} [{}]", reason, code)
            self._bus.publish(EventType.RECOVERY, {"reason": reason, "state": "record_extraction", "code": code})
        if self._debug_dir is not None:
            if result is not None and scene is not None:
                try:
                    self._record_builder.write_no_record(self._debug_dir / "no_record.json", result, scene=scene, code=code)
                except Exception as exc:
                    logger.debug("no_record write failed: {}", exc)
            else:
                self._write_debug("no_record.json", {"reason": reason, "code": code})
        self._bus.publish(EventType.NO_RECORD, {"reason": reason, "code": code})

    @staticmethod
    def _reason_code_for(reason: str) -> str:
        lowered = (reason or "").lower()
        if "no label/value pairs" in lowered:
            return "NO_PAIRS_DETECTED"
        if "all pairs empty" in lowered or "no values" in lowered:
            return "ALL_PAIRS_EMPTY"
        if reason in HARD_FAILURE_CODES:
            return reason
        return "NO_VALID_RECORD"

    def _merge_field_map(self, scene: SceneDescription) -> None:
        """Synthesise exact UIA geometry onto the observed scene when a map exists.

        The UIA field map replaces the VLM's fuzzy editable fields with exact
        controls and injects OCR source pairs, so mapping/planning/execution use
        reliable geometry even when the VLM fails to identify the form.
        """
        if self._field_map is None or not self._field_map.has_form:
            return
        origin_x, origin_y = scene.screen_offset
        added: list[ScreenElement] = []
        seen_ids: set[str] = set()

        for node in self._field_map.right_fields:
            if node.rect is None:
                continue
            element_id = f"uia-{node.handle or node.automation_id or len(added)}"
            if element_id in seen_ids:
                continue
            seen_ids.add(element_id)
            box = BBox(
                node.rect.left - origin_x,
                node.rect.top - origin_y,
                node.rect.width,
                node.rect.height,
            )
            label = (node.name or node.automation_id or "").strip()
            added.append(ScreenElement(
                element_id=element_id,
                type=node.element_type,
                label=label,
                name=node.name or "",
                bbox=box,
                confidence=1.0,
                value=None,
                required=None,
                disabled=not node.enabled,
                section="form",
                options=list(node.options),
            ))

        if self._field_map.has_source and self._ocr_callback is not None and self._field_map.left_rect is not None:
            left = self._field_map.left_rect
            try:
                lines = self._ocr_callback(BBox(left.left, left.top, left.width, left.height))
            except Exception as exc:
                logger.debug("source OCR failed: {}", exc)
                lines = []
            for label, value in pair_source_pairs(lines, self._field_map.left_labels, member_only=True):
                element_id = f"uia-src-{len(seen_ids)}"
                seen_ids.add(element_id)
                added.append(ScreenElement(
                    element_id=element_id,
                    type=ElementType.LABEL,
                    label=label,
                    name=label,
                    bbox=None,
                    confidence=0.9,
                    value=value or None,
                    section="source",
                ))

        if self._field_map.upload_button is not None and self._field_map.upload_button.rect is not None:
            btn = self._field_map.upload_button
            box = BBox(btn.rect.left - origin_x, btn.rect.top - origin_y, btn.rect.width, btn.rect.height)
            element_id = f"uia-btn-{btn.handle or 'upload'}"
            if element_id not in seen_ids:
                seen_ids.add(element_id)
                added.append(ScreenElement(
                    element_id=element_id,
                    type=ElementType.BUTTON,
                    label=btn.name or "Upload",
                    name=btn.name or "",
                    bbox=box,
                    confidence=1.0,
                    section="actions",
                ))

        if not added:
            return
        kept = [e for e in scene.elements if not e.editable]
        merged: dict[str, ScreenElement] = {e.element_id: e for e in kept}
        for element in added:
            merged[element.element_id] = element
        scene.elements = list(merged.values())
        scene.layout_summary = scene.layout_summary or "uia-anchored"

    def _find_submit(self, scene: SceneDescription) -> str | None:
        submitish = (
            "upload", "submit", "save", "next", "ok", "apply", "continue",
            "done", "finish", "update", "register", "create", "add", "confirm",
        )
        # An expandable "Upload Details" section header must never be treated as
        # the final submit button: it is clicked by _expand_upload_section so its
        # hidden fields get revealed. Any upload/attachment header that is not the
        # action-strip submit (and any already-expanded region) is excluded here
        # the same way the section expander excludes them - otherwise "upload" is
        # the first submit token and the header is wrongly clicked as submit.
        excluded: set[str] = set(self._expanded_sections)
        for candidate in find_upload_sections(scene, exclude_ids=self._expanded_sections):
            if candidate.section == "actions":
                continue
            excluded.add(candidate.element_id)
        buttons = [
            e
            for e in scene.elements
            if (e.type.value in {"button", "submit"} or e.section == "actions")
            and e.element_id not in excluded
        ]
        for label_token in submitish:
            for e in buttons:
                if label_token in (e.label or "").lower() and e.bbox is not None:
                    return e.element_id
        for e in buttons:
            if e.bbox is not None:
                return e.element_id
        for label_token in submitish:
            for e in buttons:
                if label_token in (e.label or "").lower():
                    return e.element_id
        for e in buttons:
            return e.element_id
        return None

    def _all_ok(self, results: list[ActionResult]) -> bool:
        if not results:
            return False
        return all(r.ok for r in results)

    @staticmethod
    def _skipped_fields(results: list[ActionResult]) -> list[str]:
        skipped = []
        for r in results:
            if r.success is False and r.action.field_id:
                skipped.append(r.action.field_id)
        return skipped

    @staticmethod
    def _unverified_fields(results: list[ActionResult]) -> list[str]:
        """Fields written but confirmed only as UNKNOWN (never a verified pass).

        Deduplicated by field id (the post-submit standalone VERIFY action
        re-reads the same field, so it would otherwise repeat the id).
        """
        unverified = []
        seen: set[str] = set()
        for r in results:
            if (r.success and r.verification_status == "UNKNOWN" and r.action.field_id
                    and r.action.field_id not in seen):
                unverified.append(r.action.field_id)
                seen.add(r.action.field_id)
        return unverified

    @staticmethod
    def _unmapped_required(mapping: MappingResult) -> list[str]:
        return [f.label for f in mapping.unmatched_fields if f.element.required]

    def _learn_aliases(self, record: SourceRecord, mapping: MappingResult, results: list[ActionResult]) -> None:
        """Conservatively remember fuzzy mappings that verified successfully."""
        if not self._alias_learning or self._memory is None:
            return
        verified_ids = {r.action.field_id for r in results if r.ok}
        for m in mapping.mappings:
            if m.method in {"token", "containment", "fuzzy"} and m.confidence >= 0.9 and m.target_id in verified_ids:
                try:
                    self._memory.learn_alias(m.source_label, m.target_label)
                    self._mapper.aliases.learn(m.source_label, m.target_label)
                except Exception as exc:
                    logger.debug("alias learning skipped: {}", exc)

    def _set(self, state: AgentState, detail: str | None = None) -> None:
        try:
            self._states.transition(state)
        except Exception:
            try:
                self._states.force(state)
            except Exception:
                pass
        self._state_entered[state] = time.time()
        self._state_warned.discard(state)
        if not hasattr(self, "_state_overruns"):
            self._state_overruns = {}
        if not hasattr(self, "_last_overrun_log"):
            self._last_overrun_log = {}
        self._state_overruns[state] = 0
        self._last_overrun_log.pop(state, None)
        self._bus.publish(
            EventType.STATE_CHANGED,
            {"state": self._states.state.value, "detail": detail},
        )

    # -- watchdog -------------------------------------------------------------

    @staticmethod
    def _normalize_budget(budget: float | dict[str, float] | None) -> dict[str, float]:
        if isinstance(budget, dict):
            return {k: float(v) for k, v in budget.items()}
        default = float(budget) if budget is not None else 10.0
        budgets = {state.value: default for state in AgentState}
        budgets[AgentState.WATCHING.value] = 60.0  # next-record timeout governs this
        budgets[AgentState.OBSERVING.value] = 45.0  # VLM analysis can be slow
        budgets[AgentState.THINKING.value] = 30.0
        budgets[AgentState.WAITING.value] = 60.0
        budgets[AgentState.WAITING_FOR_START_FIELD.value] = 0.0  # user-driven, never times out here
        return budgets

    def _check_state_budget(self) -> None:
        """Level-2 watchdog: surface a state that has overrun its budget.

        Level 1 is the sandbox watchdog (`ExecutionSandbox._watchdog_loop`):
        target aliveness + focus. Level 2 (here) guards the workflow state
        machine: when a state stays past its budget the overrun is counted and
        re-warned every ``_overrun_repeat_log_seconds`` while it persists, so a
        genuinely stuck state keeps surfacing instead of logging once and going
        silent. It never blocks.
        """
        state = self._states.state
        budget = self._state_budget.get(state.value, 10.0)
        if budget <= 0:
            return
        entered = self._state_entered.get(state)
        if entered is None:
            return
        elapsed = time.time() - entered
        if elapsed <= budget:
            return
        if not hasattr(self, "_state_overruns"):
            self._state_overruns = {}
        if not hasattr(self, "_last_overrun_log"):
            self._last_overrun_log = {}
        if not hasattr(self, "_overrun_repeat_log_seconds"):
            self._overrun_repeat_log_seconds = 30.0
        overruns = self._state_overruns.get(state, 0) + 1
        self._state_overruns[state] = overruns
        now = time.time()
        last_log = self._last_overrun_log.get(state, 0.0)
        if state not in self._state_warned or (now - last_log) >= self._overrun_repeat_log_seconds:
            self._state_warned.add(state)
            self._last_overrun_log[state] = now
            reason = f"state '{state.value}' overrun ({elapsed:.1f}s > {budget:.0f}s budget, tick {overruns})"
            watchdog_logger.warning("watchdog: {}", reason)
            logger.warning("watchdog: {}", reason)
            self._bus.publish(EventType.RECOVERY, {
                "reason": reason,
                "state": state.value,
                "elapsed": round(elapsed, 1),
                "budget": budget,
                "overruns": overruns,
            })


__all__ = ["AgentLoop", "RecordResult", "WorkflowSummary"]
